"""
Report routes for FlowTrack application.
Handles financial reports, exports, and analytics.
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import current_user
from src.models import db, Transaction, InitialBalance
from routes.middleware import login_required, validate_csrf_token
from src.transaction_security import get_accessible_transactions, prepare_export_data
from src.utils import calculate_totals
from src.pdf_generator import generate_cashflow_pdf
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar

report_bp = Blueprint('report', __name__)

@report_bp.route('/cashflow-statement')
@login_required
def cashflow_statement():
    """Generate cashflow statement."""
    try:
        # Get user's transactions
        transactions = get_accessible_transactions(current_user.id)
        
        if not transactions.count():
            flash('No transaction data available for cashflow statement', 'info')
            return render_template('cashflow_statement.html', statement_data=None)
        
        # Get initial balance
        initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_amount = initial_balance.amount if initial_balance else 0.0
        
        # Calculate cashflow statement data
        statement_data = calculate_cashflow_statement(transactions, initial_amount)
        
        return render_template('cashflow_statement.html', statement_data=statement_data)
        
    except Exception as e:
        flash(f'Error generating cashflow statement: {str(e)}', 'error')
        return redirect(url_for('dashboard.home'))

@report_bp.route('/generate_cashflow_statement')
@login_required
def generate_cashflow_statement():
    """Generate cashflow statement data."""
    try:
        # Get user's transactions
        transactions = get_accessible_transactions(current_user.id)
        
        if not transactions.count():
            return jsonify({'error': 'No transaction data available'}), 400
        
        # Get initial balance
        initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_amount = initial_balance.amount if initial_balance else 0.0
        
        # Calculate cashflow statement data
        statement_data = calculate_cashflow_statement(transactions, initial_amount)
        
        return jsonify({
            'success': True,
            'statement_data': statement_data
        })
        
    except Exception as e:
        return jsonify({'error': f'Error generating cashflow statement: {str(e)}'}), 500

@report_bp.route('/export_cashflow/excel')
@login_required
def export_cashflow_excel():
    """Export cashflow statement to Excel."""
    try:
        # Get user's transactions
        transactions = get_accessible_transactions(current_user.id)
        
        if not transactions.count():
            flash('No transaction data available for export', 'info')
            return redirect(url_for('report.cashflow_statement'))
        
        # Get initial balance
        initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_amount = initial_balance.amount if initial_balance else 0.0
        
        # Calculate cashflow statement data
        statement_data = calculate_cashflow_statement(transactions, initial_amount)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cashflow Statement"
        
        # Define styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        currency_format = '"$"#,##0.00'
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write title
        ws['A1'] = 'Cashflow Statement'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Write date range
        if statement_data['transactions']:
            start_date = min(t['date'] for t in statement_data['transactions'])
            end_date = max(t['date'] for t in statement_data['transactions'])
            ws['A2'] = f'Period: {start_date} to {end_date}'
            ws.merge_cells('A2:D2')
        
        # Write headers
        headers = ['Date', 'Description', 'Category', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.border = border
        
        # Write transaction data
        row = 5
        for transaction in statement_data['transactions']:
            ws.cell(row=row, column=1, value=transaction['date'])
            ws.cell(row=row, column=2, value=transaction['description'])
            ws.cell(row=row, column=3, value=transaction['category'])
            ws.cell(row=row, column=4, value=transaction['amount'])
            
            # Format amount column
            ws.cell(row=row, column=4).number_format = currency_format
            
            # Add borders
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # Write summary
        summary_row = row + 2
        ws.cell(row=summary_row, column=3, value='Total Income:').font = header_font
        ws.cell(row=summary_row, column=4, value=statement_data['total_income']).number_format = currency_format
        
        summary_row += 1
        ws.cell(row=summary_row, column=3, value='Total Expenses:').font = header_font
        ws.cell(row=summary_row, column=4, value=statement_data['total_expenses']).number_format = currency_format
        
        summary_row += 1
        ws.cell(row=summary_row, column=3, value='Net Cash Flow:').font = header_font
        ws.cell(row=summary_row, column=4, value=statement_data['net_cashflow']).number_format = currency_format
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'cashflow_statement_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        flash(f'Error exporting cashflow statement: {str(e)}', 'error')
        return redirect(url_for('report.cashflow_statement'))

@report_bp.route('/export/<file_type>')
@login_required
def export(file_type):
    """Export transactions in various formats."""
    try:
        # Get accessible transactions based on user role
        transactions = get_accessible_transactions(current_user.id)
        
        if not transactions.count():
            flash('No transaction data available for export', 'info')
            return redirect(url_for('dashboard.home'))
        
        # Prepare export data
        export_data = prepare_export_data(transactions)
        
        if file_type == 'csv':
            return export_csv(export_data)
        elif file_type == 'excel':
            return export_excel(export_data)
        elif file_type == 'pdf':
            return export_pdf(export_data)
        else:
            flash('Invalid export format', 'error')
            return redirect(url_for('dashboard.home'))
        
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'error')
        return redirect(url_for('dashboard.home'))

@report_bp.route('/forecast')
@login_required
def forecast():
    """Financial forecast page."""
    try:
        # Get user's transactions
        transactions = get_accessible_transactions(current_user.id)
        
        if not transactions.count():
            flash('No transaction data available for forecasting', 'info')
            return render_template('forecast.html', forecast_data=None)
        
        # Generate forecast data
        forecast_data = generate_forecast(transactions)
        
        return render_template('forecast.html', forecast_data=forecast_data)
        
    except Exception as e:
        flash(f'Error generating forecast: {str(e)}', 'error')
        return redirect(url_for('dashboard.home'))

@report_bp.route('/variance-analysis')
@login_required
def variance_analysis():
    """Variance analysis page."""
    try:
        # Get user's transactions
        transactions = get_accessible_transactions(current_user.id)
        
        if not transactions.count():
            flash('No transaction data available for variance analysis', 'info')
            return render_template('variance_analysis.html', analysis_data=None)
        
        # Generate variance analysis
        analysis_data = generate_variance_analysis(transactions)
        
        return render_template('variance_analysis.html', analysis_data=analysis_data)
        
    except Exception as e:
        flash(f'Error generating variance analysis: {str(e)}', 'error')
        return redirect(url_for('dashboard.home'))

@report_bp.route('/payables-aging')
@login_required
def payables_aging():
    """Payables aging report."""
    try:
        # Get user's payables (negative transactions)
        payables = get_accessible_transactions(current_user.id).filter(
            Transaction.amount < 0
        ).order_by(Transaction.date.desc()).all()
        
        # Generate aging report
        aging_data = generate_aging_report(payables, 'payables')
        
        return render_template('payables_aging.html', aging_data=aging_data)
        
    except Exception as e:
        flash(f'Error generating payables aging report: {str(e)}', 'error')
        return redirect(url_for('dashboard.home'))

@report_bp.route('/receivables-aging')
@login_required
def receivables_aging():
    """Receivables aging report."""
    try:
        # Get user's receivables (positive transactions)
        receivables = get_accessible_transactions(current_user.id).filter(
            Transaction.amount > 0
        ).order_by(Transaction.date.desc()).all()
        
        # Generate aging report
        aging_data = generate_aging_report(receivables, 'receivables')
        
        return render_template('receivables_aging.html', aging_data=aging_data)
        
    except Exception as e:
        flash(f'Error generating receivables aging report: {str(e)}', 'error')
        return redirect(url_for('dashboard.home'))

def calculate_cashflow_statement(transactions, initial_amount):
    """Calculate cashflow statement data."""
    statement_data = {
        'transactions': [],
        'total_income': 0,
        'total_expenses': 0,
        'net_cashflow': 0,
        'initial_balance': initial_amount,
        'ending_balance': initial_amount
    }
    
    for transaction in transactions.order_by(Transaction.date):
        transaction_data = {
            'date': transaction.date.isoformat(),
            'description': transaction.description,
            'category': transaction.category or 'Uncategorized',
            'amount': float(transaction.amount)
        }
        
        statement_data['transactions'].append(transaction_data)
        
        if transaction.amount > 0:
            statement_data['total_income'] += transaction.amount
        else:
            statement_data['total_expenses'] += abs(transaction.amount)
        
        statement_data['ending_balance'] += transaction.amount
    
    statement_data['net_cashflow'] = statement_data['total_income'] - statement_data['total_expenses']
    
    return statement_data

def export_csv(export_data):
    """Export data to CSV format."""
    output = BytesIO()
    
    # Create CSV content
    csv_content = "Date,Description,Category,Amount\n"
    
    for transaction in export_data:
        csv_content += f"{transaction['date']},{transaction['description']},{transaction['category']},{transaction['amount']}\n"
    
    output.write(csv_content.encode('utf-8'))
    output.seek(0)
    
    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'transactions_{datetime.now().strftime("%Y%m%d")}.csv'
    )

def export_excel(export_data):
    """Export data to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Write headers
    headers = ['Date', 'Description', 'Category', 'Amount']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row, transaction in enumerate(export_data, 2):
        ws.cell(row=row, column=1, value=transaction['date'])
        ws.cell(row=row, column=2, value=transaction['description'])
        ws.cell(row=row, column=3, value=transaction['category'])
        ws.cell(row=row, column=4, value=transaction['amount'])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'transactions_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

def export_pdf(export_data):
    """Export data to PDF format."""
    try:
        pdf_content = generate_cashflow_pdf(export_data)
        
        return send_file(
            pdf_content,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'transactions_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('dashboard.home'))

def generate_forecast(transactions):
    """Generate financial forecast."""
    # Simple forecast based on historical data
    # In a real implementation, this would use more sophisticated algorithms
    
    forecast_data = {
        'periods': [],
        'projected_income': [],
        'projected_expenses': [],
        'projected_balance': []
    }
    
    # Calculate average monthly income and expenses
    monthly_data = {}
    
    for transaction in transactions:
        month_key = transaction.date.strftime('%Y-%m')
        
        if month_key not in monthly_data:
            monthly_data[month_key] = {'income': 0, 'expenses': 0}
        
        if transaction.amount > 0:
            monthly_data[month_key]['income'] += transaction.amount
        else:
            monthly_data[month_key]['expenses'] += abs(transaction.amount)
    
    if monthly_data:
        avg_income = sum(data['income'] for data in monthly_data.values()) / len(monthly_data)
        avg_expenses = sum(data['expenses'] for data in monthly_data.values()) / len(monthly_data)
        
        # Generate 12-month forecast
        current_date = datetime.now()
        current_balance = sum(t.amount for t in transactions)
        
        for i in range(12):
            forecast_date = current_date + timedelta(days=30 * i)
            forecast_data['periods'].append(forecast_date.strftime('%Y-%m'))
            forecast_data['projected_income'].append(avg_income)
            forecast_data['projected_expenses'].append(avg_expenses)
            
            current_balance += avg_income - avg_expenses
            forecast_data['projected_balance'].append(current_balance)
    
    return forecast_data

def generate_variance_analysis(transactions):
    """Generate variance analysis."""
    # Simple variance analysis comparing actual vs expected
    # In a real implementation, this would compare against budgets or forecasts
    
    analysis_data = {
        'categories': {},
        'total_variance': 0
    }
    
    # Group transactions by category
    category_totals = {}
    
    for transaction in transactions:
        category = transaction.category or 'Uncategorized'
        
        if category not in category_totals:
            category_totals[category] = 0
        
        category_totals[category] += transaction.amount
    
    # Calculate variance (simplified - comparing to zero baseline)
    for category, total in category_totals.items():
        analysis_data['categories'][category] = {
            'actual': total,
            'expected': 0,  # In real implementation, this would come from budget
            'variance': total,
            'variance_percent': 100 if total != 0 else 0
        }
    
    analysis_data['total_variance'] = sum(t.amount for t in transactions)
    
    return analysis_data

def generate_aging_report(transactions, report_type):
    """Generate aging report for payables or receivables."""
    aging_data = {
        'current': 0,
        '30_days': 0,
        '60_days': 0,
        '90_days': 0,
        'over_90_days': 0,
        'transactions': []
    }
    
    current_date = datetime.now().date()
    
    for transaction in transactions:
        days_old = (current_date - transaction.date).days
        
        if days_old <= 30:
            aging_data['current'] += abs(transaction.amount)
        elif days_old <= 60:
            aging_data['30_days'] += abs(transaction.amount)
        elif days_old <= 90:
            aging_data['60_days'] += abs(transaction.amount)
        elif days_old <= 120:
            aging_data['90_days'] += abs(transaction.amount)
        else:
            aging_data['over_90_days'] += abs(transaction.amount)
        
        aging_data['transactions'].append({
            'date': transaction.date.isoformat(),
            'description': transaction.description,
            'amount': abs(transaction.amount),
            'days_old': days_old
        })
    
    return aging_data
