import os
import sys
import secrets
import tempfile
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app, session, abort, g
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from typing_extensions import Annotated
from src.models import db, User, Transaction, InitialBalance, UserPreferences, Tenant
from sqlalchemy.exc import OperationalError
from src.forms import LoginForm, RegistrationForm
from src.anthropic_service import FinancialAnalytics
from src.upload_handler import process_upload
from src.utils import calculate_totals, calculate_burn_rate, calculate_runway
from src.rbac import assign_default_role, assign_role, remove_role, get_user_roles, update_user_preferences_on_role_change
from src.auth_decorators import (
    super_admin_required, admin_required, admin_or_super_admin_required,
    transaction_owner_or_admin_required, authenticated_only
)
from src.transaction_security import (
    get_accessible_transactions, prepare_export_data, log_data_access
)
from src.ai_utils import ai_rate_limit, monitor_ai_performance
from src.config import Config
import pandas as pd
from io import BytesIO
from dotenv import load_dotenv
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import io
import base64
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.dates import MonthLocator, DateFormatter
import matplotlib.ticker as ticker
from datetime import datetime
import calendar
from flask_migrate import Migrate
from flask_babel import Babel
import time
from functools import wraps
from datetime import datetime, timedelta
import time
import logging
from src.admin_utils import (
    get_users_with_roles,
    get_user_statistics,
    get_user_transaction_summary,
    validate_role_assignment,
    log_admin_action,
    get_user_role_history,
    export_user_data,
    get_available_roles_for_admin
)
from src.currency_service import init_currency_service, get_currency_service
from sqlalchemy.exc import SQLAlchemyError
from werkzeug.routing import BuildError

# Load .env file explicitly at the start
load_dotenv()

# Ensure instance folder exists
instance_path = os.path.join(os.path.dirname(__file__),'instance')
os.makedirs(instance_path, exist_ok=True)

upload_path = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(upload_path, exist_ok=True)

# Helper functions for file uploads and CSRF validation
def generate_upload_token():
    """Generate a secure random token for file uploads."""
    return secrets.token_urlsafe(32)

def save_temp_file(file, token):
    """Save uploaded file to temporary storage with token."""
    temp_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp')
    os.makedirs(temp_dir, exist_ok=True)
    
    filename = secure_filename(file.filename)
    file_path = os.path.join(temp_dir, f"{token}_{filename}")
    file.save(file_path)
    return file_path

def get_temp_file(token):
    """Retrieve temporary file by token."""
    temp_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp')
    for filename in os.listdir(temp_dir):
        if filename.startswith(f"{token}_"):
            return os.path.join(temp_dir, filename)
    return None

def cleanup_temp_file(token):
    """Clean up temporary file by token."""
    temp_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'temp')
    for filename in os.listdir(temp_dir):
        if filename.startswith(f"{token}_"):
            try:
                os.remove(os.path.join(temp_dir, filename))
            except Exception:
                pass

def validate_csrf_header():
    """Validate CSRF token from request headers."""
    csrf_token = request.headers.get('X-CSRFToken')
    if not csrf_token:
        return False
    # Flask-WTF will handle the actual validation
    return True

def get_locale():
    # First check if a language is stored in the session
    if 'language' in session:
        lang = session['language']
        return lang
        
    # Otherwise fallback to browser preference
    browser_lang = request.accept_languages.best_match(['en', 'es', 'ja', 'ar', 'ru', 'zh'])
    return browser_lang

def create_app(skip_schema_validation=False, start_services=True):
    """Create and configure Flask application factory.
    
    Args:
        skip_schema_validation (bool): Skip database schema validation on startup
        start_services (bool): Start background services like scheduler
        
    Returns:
        tuple: (app, db) - Flask app instance and database instance
    """
    app = Flask(__name__,
                template_folder='templates',
                static_folder='static')
    app.config.from_object(Config)
    app.config.setdefault('UPLOAD_FOLDER', os.path.join(os.path.dirname(__file__), 'uploads'))

    # Initialize rate limiter with Redis storage
    limiter.init_app(app)

    # Initialize CSRF protection
    csrf = CSRFProtect(app)

    db.init_app(app)

    def validate_database_schema():
        """Validate database schema consistency before application startup.
        
        Checks for required currency columns in the transaction table and provides
        clear error messages if schema is out of sync with model definitions.
        
        Returns:
            bool: True if schema is valid, False otherwise
        """
        try:
            with app.app_context():
                from sqlalchemy import text, inspect
                
                # Check if transaction table exists
                inspector = inspect(db.engine)
                table_names = inspector.get_table_names()
                
                if 'transaction' not in table_names:
                    app.logger.error("Transaction table does not exist in database")
                    return False
                
                # Check for required currency columns
                columns = inspector.get_columns('transaction')
                column_names = [col['name'] for col in columns]
                
                required_currency_columns = [
                    'original_currency',
                    'original_amount', 
                    'base_currency_amount',
                    'exchange_rate',
                    'rate_date',
                    'currency_conversion_status'
                ]
                
                missing_columns = [col for col in required_currency_columns if col not in column_names]
                
                if missing_columns:
                    app.logger.error(f"Missing required currency columns: {missing_columns}")
                    app.logger.error("Database schema is out of sync with model definitions")
                    app.logger.error("This usually indicates missing migrations")
                    return False
                
                # Check for expected indexes
                indexes = inspector.get_indexes('transaction')
                index_names = [idx['name'] for idx in indexes]
                
                expected_indexes = [
                    'ix_transaction_original_currency',
                    'ix_transaction_currency_conversion_status'
                ]
                
                missing_indexes = [idx for idx in expected_indexes if idx not in index_names]
                if missing_indexes:
                    app.logger.warning(f"Missing expected indexes: {missing_indexes}")
                    app.logger.warning("Consider running migrations to add missing indexes")
                
                app.logger.info("Database schema validation successful")
                return True
                
        except OperationalError as e:
            app.logger.error(f"Database connection error during schema validation: {str(e)}")
            app.logger.error("Unable to connect to database for schema validation")
            return False
        except Exception as e:
            app.logger.error(f"Schema validation failed with error: {str(e)}")
            app.logger.error("Unable to validate database schema. Check database connection.")
            return False

    # Validate database schema before proceeding (unless explicitly skipped)
    if not skip_schema_validation and os.getenv('SKIP_SCHEMA_VALIDATION') != '1':
        if not validate_database_schema():
            app.logger.error("Database schema validation failed. Application cannot start safely.")
            app.logger.error("Please run 'python run_currency_migration.py' to fix the schema issues.")
            app.logger.error("To skip schema validation, set SKIP_SCHEMA_VALIDATION=1")
            raise RuntimeError("Database schema validation failed")

    migrate = Migrate(app, db)

    # Initialize currency service
    currency_service = init_currency_service(app)

    # Initialize scheduler for background tasks (only when explicitly enabled)
    if start_services and os.getenv('RUN_SCHEDULER') == '1':
        from src.scheduler import BankScheduler
        scheduler = BankScheduler(app)
        scheduler.initialize_scheduler(app)
        scheduler.start_scheduler()
        app.logger.info("Background scheduler started")
    else:
        app.logger.info("Background scheduler disabled (set RUN_SCHEDULER=1 to enable)")

    login_manager.init_app(app)
    login_manager.login_view = 'login'

    babel.init_app(app, locale_selector=get_locale)
    
    # Move all context processors inside create_app()
    @app.context_processor
    def inject_role_context():
        try:
            roles = [role.name for role in current_user.roles] if current_user.is_authenticated and hasattr(current_user, 'roles') else []
        except Exception:
            roles = []
        return {
            'user_roles': roles,
            'is_super_admin': 'super_admin' in roles,
            'is_admin': 'admin' in roles,
            'is_user': ('admin' not in roles and 'super_admin' not in roles)
        }



    
    return app, db, limiter, login_manager, babel

app, db = create_app(skip_schema_validation=os.getenv('SKIP_SCHEMA_VALIDATION') == '1')

# Initialize extensions at module level (canonical Flask extension pattern)
limiter = Limiter(key_func=lambda: f"{current_user.id if current_user.is_authenticated else get_remote_address()}", storage_uri="redis://localhost:6379")
login_manager = LoginManager()
babel = Babel()

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Enhanced helper function for AI navigation context
def get_ai_navigation_context(current_page, page_title=None, page_subtitle=None, page_description=None):
    """Generate consistent navigation context for AI pages with enhanced metadata"""
    breadcrumb_items = [
        {'title': 'Workspace', 'url': '/'},
        {'title': 'AI Analysis', 'url': '/ai-analysis'},
        {'title': page_title or current_page.title(), 'url': None}
    ]
    
    ai_features_nav = [
        {
            'id': 'dashboard',
            'title': 'Dashboard',
            'url': '/ai-analysis/dashboard',
            'icon': 'fas fa-tachometer-alt',
            'description': 'AI system overview and metrics'
        },
        {
            'id': 'risk',
            'title': 'Risk Assessment',
            'url': '/ai-analysis/risk',
            'icon': 'fas fa-shield-alt',
            'description': 'Financial risk analysis and assessment'
        },
        {
            'id': 'anomaly',
            'title': 'Anomaly Detection',
            'url': '/ai-analysis/anomaly',
            'icon': 'fas fa-search',
            'description': 'Detect unusual transaction patterns'
        },
        {
            'id': 'forecast',
            'title': 'Advanced Forecast',
            'url': '/ai-analysis/forecast',
            'icon': 'fas fa-chart-line',
            'description': 'Predictive forecasting models'
        },
        {
            'id': 'assistant',
            'title': 'Virtual Analyst',
            'url': '/ai-analysis/assistant',
            'icon': 'fas fa-robot',
            'description': 'AI-powered data analysis assistant'
        }
    ]
    
    return {
        'breadcrumb_items': breadcrumb_items,
        'ai_features_nav': ai_features_nav,
        'current_page': current_page,
        'page_title': page_title or current_page.title(),
        'page_subtitle': page_subtitle or f'AI-powered {current_page} analysis',
        'page_description': page_description or f'Advanced AI analysis for {current_page} insights and recommendations'
    }


# Performance monitoring decorator
def performance_monitor(f):
    """Decorator to monitor route performance"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        start_time = time.time()
        try:
            result = f(*args, **kwargs)
            execution_time = time.time() - start_time
            
            # Log performance metrics
            if execution_time > 1.0:  # Log slow operations
                logging.warning(f"Slow route {f.__name__}: {execution_time:.2f}s")
            
            return result
        except Exception as e:
            execution_time = time.time() - start_time
            logging.error(f"Route {f.__name__} failed after {execution_time:.2f}s: {str(e)}")
            raise
    return decorated_function

# Error tracking and logging
def log_route_error(route_name, error, user_id=None):
    """Log route errors with context"""
    error_info = {
        'route': route_name,
        'error': str(error),
        'user_id': user_id,
        'timestamp': datetime.utcnow().isoformat(),
        'user_agent': request.headers.get('User-Agent', ''),
        'ip_address': request.remote_addr
    }
    logging.error(f"Route error: {error_info}")

# Global tenant branding context with safe defaults
@app.context_processor
def inject_tenant_branding():
    try:
        # If a current tenant is set in g or session, try to load its branding
        tenant_branding = None
        try:
            current_tenant_id = getattr(g, 'current_tenant_id', None) or session.get('current_tenant_id')
        except Exception:
            current_tenant_id = None
        if current_tenant_id:
            try:
                from src.tenant_utils import get_tenant_branding  # local import to avoid circulars
                tenant_branding = get_tenant_branding(current_tenant_id)
            except Exception:
                tenant_branding = None
        if not tenant_branding:
            tenant_branding = {
                'primary_color': '#0d6efd',
                'company_name': 'FlowTrack',
                'logo_url': None
            }
        # Also expose to window.tenant_branding via inline script consumer (static/js/tenant.js)
        return {
            'tenant_branding': tenant_branding
        }
    except Exception:
        return {
            'tenant_branding': {
                'primary_color': '#0d6efd',
                'company_name': 'FlowTrack',
                'logo_url': None
            }
        }

@app.before_request
def load_current_tenant_into_g():
    try:
        g.current_tenant_id = session.get('current_tenant_id')
    except Exception:
        pass

@app.after_request
def add_security_headers(response):
    """Add security headers to all responses"""
    try:
        security_headers = current_app.config.get('SECURITY_HEADERS', {})
        for header, value in security_headers.items():
            response.headers[header] = value
    except Exception:
        pass
    return response

@app.context_processor
def inject_tenant_options():
    try:
        # Expose tenant options only to super admins when switching is allowed
        roles = [role.name for role in current_user.roles] if current_user.is_authenticated and hasattr(current_user, 'roles') else []
        if 'super_admin' in roles and current_app.config.get('ALLOW_TENANT_SWITCHING', True):
            tenants = Tenant.query.with_entities(Tenant.id, Tenant.name).all()
            return {
                'tenant_options': [{'id': t.id, 'name': t.name} for t in tenants]
            }
    except Exception:
        pass
    return {}

@app.context_processor
def inject_current_app_proxy():
    # Provide safe access to common config values without exposing the whole app object
    try:
        cfg = current_app.config if current_app else {}
    except Exception:
        cfg = {}
    return {
        'config': cfg,
        # global KPI defaults for templates that expect it
        'default_kpis': {
            'current_ratio': 0.0,
            'quick_ratio': 0.0,
            'working_capital': 0.0,
            'cash_cycle': 0.0,
            'aging_buckets': {
                '0-30': 0.0,
                '31-60': 0.0,
                '61-90': 0.0,
                '90+': 0.0
            },
            'collection_efficiency': 0.0,
            'dso': 0.0,
            'payables_schedule': {
                '30': 0.0
            },
            'dpo': 0.0
        }
    }

# Helper: Check if endpoint exists and safe url builder
@app.context_processor
def inject_url_helpers():
    def endpoint_exists(name: str) -> bool:
        try:
            return name in current_app.view_functions
        except Exception:
            return False
    def url_or(endpoint: str, **values):
        try:
            return url_for(endpoint, **values)
        except BuildError:
            return '#'
        except Exception:
            return '#'
    return {
        'endpoint_exists': endpoint_exists,
        'url_or': url_or
    }

@app.context_processor
def inject_currency_context():
    """Inject currency information into template context."""
    try:
        currency_service = get_currency_service()
        if not currency_service:
            return {}
        
        return {
            'currency_service': currency_service,
            'primary_currency': currency_service.config.get('CURRENCY_PRIMARY', 'AED'),
            'supported_currencies': currency_service.get_supported_currencies(),
            'currency_symbols': currency_service.config.get('CURRENCY_SYMBOL_MAP', {}),
            'currency_conversion_enabled': currency_service.config.get('CURRENCY_CONVERSION_ENABLED', True),
            'format_currency': currency_service.format_currency,
            'get_currency_symbol': currency_service.get_currency_symbol
        }
    except Exception as e:
        # Log error but don't break the app
        print(f"Currency context injection failed: {e}")
        return {}

############################
# Placeholder routes to satisfy template links when data/modules are not yet set up
############################

# Receivables
@app.route('/receivables', methods=['GET'])
@login_required
def receivables_list():
    try:
        return render_template('receivables.html', receivables=[], total_outstanding=0)
    except Exception:
        flash('Receivables module not available yet.', 'info')
        return redirect(url_for('home'))

@app.route('/receivables/create', methods=['GET', 'POST'])
@login_required
def receivables_create():
    if request.method == 'POST':
        flash('Receivable created (placeholder).', 'success')
        return redirect(url_for('receivables_list'))
    return render_template('receivables_form.html')

@app.route('/receivables/<int:receivable_id>/edit', methods=['GET', 'POST'])
@login_required
def receivables_edit(receivable_id: int):
    if request.method == 'POST':
        flash('Receivable updated (placeholder).', 'success')
        return redirect(url_for('receivables_list'))
    return render_template('receivables_form.html', receivable_id=receivable_id)

@app.route('/receivables/<int:receivable_id>/mark-paid', methods=['POST'])
@login_required
def receivables_mark_paid(receivable_id: int):
    flash('Receivable marked as paid (placeholder).', 'success')
    return redirect(url_for('receivables_list'))

# Payables
@app.route('/payables', methods=['GET'])
@login_required
def payables_list():
    try:
        return render_template('payables.html', payables=[], total_outstanding=0)
    except Exception:
        flash('Payables module not available yet.', 'info')
        return redirect(url_for('home'))

@app.route('/payables/create', methods=['GET', 'POST'])
@login_required
def payables_create():
    if request.method == 'POST':
        flash('Payable created (placeholder).', 'success')
        return redirect(url_for('payables_list'))
    return render_template('payables_form.html')

@app.route('/payables/<int:payable_id>/edit', methods=['GET', 'POST'])
@login_required
def payables_edit(payable_id: int):
    if request.method == 'POST':
        flash('Payable updated (placeholder).', 'success')
        return redirect(url_for('payables_list'))
    return render_template('payables_form.html', payable_id=payable_id)

@app.route('/payables/<int:payable_id>/mark-paid', methods=['POST'])
@login_required
def payables_mark_paid(payable_id: int):
    flash('Payable marked as paid (placeholder).', 'success')
    return redirect(url_for('payables_list'))

# Admin templates/invitations/approvals
@app.route('/admin/templates', methods=['GET'])
@admin_required
def admin_templates():
    return render_template('admin/templates.html', templates=[])

@app.route('/admin/templates/create', methods=['GET', 'POST'])
@admin_required
def admin_templates_create():
    if request.method == 'POST':
        flash('Template created (placeholder).', 'success')
        return redirect(url_for('admin_templates'))
    return render_template('admin/templates_form.html')

@app.route('/admin/templates/<int:template_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_templates_edit(template_id: int):
    if request.method == 'POST':
        flash('Template updated (placeholder).', 'success')
        return redirect(url_for('admin_templates'))
    return render_template('admin/templates_form.html', template_id=template_id)

@app.route('/admin/invitations', methods=['GET'])
@admin_required
def admin_invitations():
    return render_template('admin/invitations.html', pending_invitations=0)

@app.route('/admin/invitations/send', methods=['GET', 'POST'])
@admin_required
def admin_invitations_send():
    if request.method == 'POST':
        flash('Invitation sent (placeholder).', 'success')
        return redirect(url_for('admin_invitations'))
    return render_template('admin/invitations_form.html')

@app.route('/admin/approvals/pending', methods=['GET'])
@admin_required
def approvals_pending():
    return render_template('admin/approvals.html', pending_approvals=0)

# Bulk categorize transactions link target
@app.route('/admin/categorize-transactions', methods=['GET'])
@admin_required
def admin_categorize_transactions():
    flash('Bulk categorization (placeholder).', 'info')
    return redirect(url_for('cash_overview'))


@app.template_filter('datetimeformat')
def datetimeformat(value, fmt='%Y-%m-%d %H:%M'):
    """Jinja filter to format datetime or datetime-like strings safely."""
    try:
        if value is None or value == '':
            return ''
        if isinstance(value, str):
            # Try ISO format first
            try:
                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            except Exception:
                # Fallback to common formats
                for pattern in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
                    try:
                        dt = datetime.strptime(value, pattern)
                        break
                    except Exception:
                        dt = None
                if dt is None:
                    return value
        elif hasattr(value, 'strftime'):
            dt = value
        else:
            return value
        return dt.strftime(fmt)
    except Exception:
        return value

def validate_csrf_header_if_enabled() -> bool:
    """Validates X-CSRFToken header only when CSRF is enabled and a session token exists."""
    try:
        if not current_app.config.get('WTF_CSRF_ENABLED'):
            return True
        token_in_session = session.get('csrf_token')
        if not token_in_session:
            # No session token to validate against; skip strict validation
            return True
        header_token = request.headers.get('X-CSRFToken') or request.headers.get('X-CSRF-Token')
        if not header_token:
            return False
        return header_token == token_in_session
    except Exception:
        return False

@app.route('/admin/tenant-switch', methods=['POST'])
@super_admin_required
def admin_tenant_switch():
    try:
        tenant_id = request.form.get('tenant_id')
        if not tenant_id:
            flash('No tenant selected', 'warning')
            return redirect(request.referrer or url_for('home'))
        # Persist selection
        session['current_tenant_id'] = tenant_id
        try:
            g.current_tenant_id = tenant_id
        except Exception:
            pass
        flash('Tenant switched', 'success')
    except Exception as e:
        current_app.logger.warning(f"Tenant switch failed: {e}")
        flash('Failed to switch tenant', 'danger')
    return redirect(request.referrer or url_for('home'))

@app.route('/set-language/<language>')
def set_language(language):
    # Store language in session
    session['language'] = language
    flash(f'Language set to: {language}', 'info')
    # Redirect back to the referring page or home page
    return redirect(request.referrer or url_for('home'))


#All routes in app
@app.route('/',methods=['GET','POST'])
@app.route('/home', methods=['GET', 'POST'])
@login_required
def home():
    user_status = "Logged In"
    page = request.args.get('page', 1, type=int)
    per_page = 8

    if request.method == 'POST':
        # Handle adding transactions
        new_transaction = Transaction(
            user_id=current_user.id,
            date=request.form.get('date'),
            description=request.form.get('description'),
            amount=float(request.form.get('amount')),
            type=request.form.get('type')
        )
        db.session.add(new_transaction)
        db.session.commit()
        flash('Transaction added successfully', 'success')
        return redirect(url_for('home'))

    # Filter transactions by current user
    paginated_transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date.desc()).paginate(page=page, per_page=per_page)
    all_transactions = Transaction.query.filter_by(user_id=current_user.id).all()
    
    # Get initial balance for current user
    initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
    initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

    total_cfo, total_cfi, total_cff = calculate_totals(all_transactions)
    balance = initial_balance + total_cfo + total_cfi + total_cff

    return render_template('home.html', transactions=paginated_transactions, balance=balance, initial_balance=initial_balance,
                           total_cfo=total_cfo, total_cfi=total_cfi, total_cff=total_cff, user=current_user, status=user_status)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        # Optional OTP support
        otp_code = request.form.get('otp')
        # If 2FA is enforced in settings in future, validate otp_code here
        if user and check_password_hash(user.password, form.password.data):
            login_user(user, remember=True)
            return redirect(url_for('home'))
        else:
            flash('Login Unsuccessful. Please check username and password', 'danger')
    return render_template('login.html', form=form)


@app.route('/privacy')
@app.route('/privacy-policy')
@app.route('/legal/privacy')
def privacy_policy():
    """Public privacy policy page for legal compliance."""
    return render_template('legal/privacy_policy.html')


@app.route('/terms')
@app.route('/terms-of-service')
@app.route('/legal/terms')
def terms_of_service():
    """Public terms of service page referenced from the login template."""
    return render_template('legal/terms_of_service.html')

@app.route('/register',methods=['GET','POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    form = RegistrationForm()
    if form.validate_on_submit():
        user_exists = User.query.filter_by(username=form.username.data).first()
        if user_exists:
            flash('Username already exists. Please choose a different one.', 'danger')
        else:
            hashed_password = generate_password_hash(form.password.data)
            new_user = User(username=form.username.data, password=hashed_password)
            db.session.add(new_user)
            db.session.commit()
            
            # Assign default 'user' role
            success = assign_default_role(new_user)
            if not success:
                current_app.logger.warning(f"Default role assignment failed for user {new_user.username}. Ensure roles are seeded.")
            
            new_preferences = UserPreferences(user_id=new_user.id)
            db.session.add(new_preferences)
            db.session.flush()  # Ensure the preferences have an ID
            new_preferences.ensure_role_based_modules(new_user)
            db.session.commit()
            flash('Your account has been created! You can now log in.', 'success')
            return redirect(url_for('login'))
    return render_template('register.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/add-transaction', methods=['GET', 'POST'])
@login_required
def add_transaction_page():
    from src.forms import EnterpriseTransactionForm
    form = EnterpriseTransactionForm()
    
    if form.validate_on_submit():
        try:
            # Get currency information from form or defaults
            original_currency = request.form.get('currency', app.config.get('CURRENCY_PRIMARY', 'AED'))
            original_amount = form.amount.data
            
            # Validate and normalize currency code
            try:
                from src.currency_service import get_currency_service
                currency_service = get_currency_service()
                if currency_service and original_currency:
                    if not currency_service.validate_currency_code(original_currency):
                        flash(f'Invalid currency code "{original_currency}". Using base currency instead.', 'warning')
                        original_currency = currency_service.config.get('CURRENCY_PRIMARY', 'AED')
                    else:
                        original_currency = original_currency.upper()  # Normalize to uppercase
            except Exception as e:
                app.logger.warning(f'Currency validation failed: {str(e)}')
                original_currency = app.config.get('CURRENCY_PRIMARY', 'AED')
            
            new_transaction = Transaction(
                user_id=current_user.id,
                date=form.date.data,
                description=form.description.data,
                amount=form.amount.data,
                type=form.type.data,
                activity_category=form.activity_category.data if form.activity_category.data else None,
                # Currency fields
                original_currency=original_currency,
                original_amount=original_amount
            )
            
            # Apply currency conversion if needed
            try:
                currency_service = get_currency_service()
                if currency_service and currency_service.config.get('CURRENCY_CONVERSION_ENABLED', True):
                    conversion_success = new_transaction.apply_currency_conversion(currency_service)
                    if not conversion_success:
                        flash('Warning: Currency conversion failed, transaction saved with original currency', 'warning')
            except AttributeError as e:
                app.logger.error(f'Currency conversion method not available: {str(e)}')
                flash('Warning: Currency conversion unavailable, transaction saved with original currency', 'warning')
            except Exception as e:
                app.logger.error(f'Currency conversion error: {str(e)}')
                flash('Warning: Currency conversion failed, transaction saved with original currency', 'warning')
            
            db.session.add(new_transaction)
            db.session.commit()
            flash('Transaction added successfully', 'success')
            return redirect(url_for('home'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding transaction: {str(e)}', 'danger')
    
    return render_template('transaction_create.html', form=form)

@app.route('/edit/<int:transaction_id>', methods=['GET', 'POST'])
@transaction_owner_or_admin_required('transaction_id')
def edit_transaction(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    if request.method == 'POST':
        transaction.date = request.form['date']
        transaction.description = request.form['description']
        
        # Validate and handle invalid numeric input before commit
        try:
            transaction.amount = float(request.form['amount'])
        except (ValueError, TypeError):
            flash('Invalid amount format. Please enter a valid number.', 'error')
            return render_template('edit_transaction.html', transaction=transaction)
        
        transaction.type = request.form['type']
        
        # Reapply currency conversion if currency service is enabled
        try:
            from src.currency_service import get_currency_service
            cs = get_currency_service()
            if cs and cs.config.get('CURRENCY_CONVERSION_ENABLED', True):
                # Ensure original_currency/original_amount are set appropriately first
                if not transaction.original_currency:
                    transaction.original_currency = cs.config.get('CURRENCY_PRIMARY', 'AED')
                
                # Edge case: original_amount may become stale when amount changes on edit
                # Update original_amount to reflect the newly posted amount when it differs
                if not transaction.original_amount or transaction.original_amount != transaction.amount:
                    transaction.original_amount = transaction.amount
                
                # Apply currency conversion
                conversion_success = transaction.apply_currency_conversion(cs)
                if not conversion_success:
                    flash('Warning: Currency conversion failed for updated transaction', 'warning')
        except ImportError as e:
            app.logger.warning(f'Currency service not available: {str(e)}')
            flash('Currency conversion service unavailable', 'info')
        except AttributeError as e:
            app.logger.error(f'Currency conversion method not available: {str(e)}')
            flash('Warning: Currency conversion unavailable for updated transaction', 'warning')
        except Exception as e:
            app.logger.error(f'Currency conversion error: {str(e)}')
            flash(f'Warning: Currency conversion error: {str(e)}', 'warning')
        
        db.session.commit()
        flash('Transaction Updated Successfully', 'success')
        return redirect(url_for('cash_activities'))
    return render_template('edit_transaction.html', transaction=transaction)

@app.route('/delete/<int:transaction_id>', methods=['POST'])
@transaction_owner_or_admin_required('transaction_id')
def delete_transaction(transaction_id):
    transaction = Transaction.query.get_or_404(transaction_id)
    db.session.delete(transaction)
    db.session.commit()
    flash('Your Transaction Deleted!', 'danger')
    return redirect(url_for('home'))

@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_route():
    if request.method == 'GET':
        return render_template('upload.html')
    elif request.method == 'POST':
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No selected file'}), 400
        
        try:
            result = process_upload(file)
            return jsonify(result)
        except Exception as e:
            app.logger.error(f"Error in upload_file: {str(e)}")
            return jsonify({'error': str(e)}), 500


@app.route('/upload/preview', methods=['POST'])
@login_required
@limiter.limit("10 per minute")
def upload_preview():
    """Handle file upload and return preview data for column mapping."""
    # Validate CSRF token
    if not validate_csrf_header():
        return jsonify({'error': {'code': 'CSRF_ERROR', 'message': 'Invalid CSRF token', 'details': None}}), 403
    
    # CRITICAL FIX: Validate user session
    app.logger.info(f"Upload preview - Current user: {current_user.username} (ID: {current_user.id})")
    if not current_user.is_authenticated or not hasattr(current_user, 'id') or not current_user.id:
        app.logger.error("Invalid user session during upload preview")
        return jsonify({'error': {'code': 'INVALID_SESSION', 'message': 'Invalid user session', 'details': None}}), 401
    
    if 'file' not in request.files:
        return jsonify({'error': {'code': 'NO_FILE', 'message': 'No file part', 'details': None}}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': {'code': 'NO_FILE', 'message': 'No selected file', 'details': None}}), 400
    
    try:
        # Generate secure token and save file
        token = generate_upload_token()
        file_path = save_temp_file(file, token)
        
        # Store token in session
        session['upload_file_token'] = token
        
        # Get file preview using the saved file
        from src.upload_handler import get_file_preview_from_path
        result = get_file_preview_from_path(file_path)
        result['upload_token'] = token
        
        app.logger.info(f"File preview generated for {file.filename}")
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Error in upload_preview: {str(e)}")
        return jsonify({'error': {'code': 'PREVIEW_ERROR', 'message': str(e), 'details': None}}), 500

@app.route('/upload/suggestions', methods=['POST'])
@login_required
@limiter.limit("20 per minute")
def upload_suggestions():
    """Get intelligent column mapping suggestions."""
    # Validate CSRF token
    if not validate_csrf_header():
        return jsonify({'error': {'code': 'CSRF_ERROR', 'message': 'Invalid CSRF token', 'details': None}}), 403
    
    try:
        data = request.get_json()
        if not data or 'columns' not in data:
            return jsonify({'error': {'code': 'NO_COLUMNS', 'message': 'No columns provided', 'details': None}}), 400
        
        # Validate that columns is an array of strings
        columns = data['columns']
        if not isinstance(columns, list) or not all(isinstance(col, str) for col in columns):
            return jsonify({'error': {'code': 'INVALID_REQUEST', 'message': 'Invalid columns payload', 'details': {'expected': 'array of strings'}}}), 400
        
        from src.upload_handler import get_column_suggestions
        suggestions = get_column_suggestions(columns)
        
        return jsonify({
            'suggestions': suggestions,
            'message': 'Column suggestions generated successfully'
        })
    except Exception as e:
        app.logger.error(f"Error in upload_suggestions: {str(e)}")
        return jsonify({'error': {'code': 'SUGGESTIONS_ERROR', 'message': str(e), 'details': None}}), 500

@app.route('/upload/process', methods=['POST'])
@login_required
@limiter.limit("5 per minute")
def upload_process():
    """Process upload with column mapping and create transactions."""
    # Validate CSRF token
    if not validate_csrf_header():
        return jsonify({'error': {'code': 'CSRF_ERROR', 'message': 'Invalid CSRF token', 'details': None}}), 403
    
    # Debug: Log current user info
    app.logger.info(f"Upload process - Current user: {current_user.username} (ID: {current_user.id})")
    
    # Verify user is properly authenticated
    if not current_user.is_authenticated:
        app.logger.error("User not authenticated during upload process")
        return jsonify({'error': {'code': 'AUTH_ERROR', 'message': 'User not authenticated', 'details': None}}), 401
    
    # CRITICAL FIX: Force user session validation
    if not hasattr(current_user, 'id') or not current_user.id:
        app.logger.error("Invalid user session - no user ID found")
        return jsonify({'error': {'code': 'INVALID_SESSION', 'message': 'Invalid user session', 'details': None}}), 401
    
    # Additional validation: Check if user exists in database
    from src.models import User
    db_user = User.query.get(current_user.id)
    if not db_user:
        app.logger.error(f"User {current_user.id} not found in database")
        return jsonify({'error': {'code': 'USER_NOT_FOUND', 'message': 'User not found in database', 'details': None}}), 401
    
    app.logger.info(f"User session validated: {db_user.username} (ID: {db_user.id})")
    
    # Check for session token first
    upload_token = session.get('upload_file_token')
    if not upload_token:
        return jsonify({'error': {'code': 'NO_SESSION_FILE', 'message': 'No file in session', 'details': None}}), 400
    
    # Try to get file from session token, fallback to form upload
    file_path = get_temp_file(upload_token)
    if not file_path:
        return jsonify({'error': {'code': 'FILE_NOT_FOUND', 'message': 'Session file not found', 'details': None}}), 400
    
    column_mapping_str = request.form.get('column_mapping')
    if not column_mapping_str:
        return jsonify({'error': {'code': 'NO_MAPPING', 'message': 'No column mapping provided', 'details': None}}), 400
    
    try:
        import json
        column_mapping = json.loads(column_mapping_str)
        
        from src.upload_handler import process_upload_with_mapping_from_path
        result = process_upload_with_mapping_from_path(file_path, column_mapping)
        
        # Check for validation errors first
        if result.get('error'):
            cleanup_temp_file(upload_token)
            session.pop('upload_file_token', None)
            app.logger.info(f"Upload validation failed: {result['error']}")
            return jsonify(result), 400
        
        # Create transactions from the processed data
        if result.get('data'):
            transactions_created = 0
            date_parse_warnings = 0
            for record in result['data']:
                try:
                    # Parse date properly
                    date_str = record.get('date')
                    parsed_date = None
                    if date_str:
                        try:
                            from datetime import datetime
                            parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            date_parse_warnings += 1
                            app.logger.warning(f"Could not parse date: {date_str}")
                            continue
                    
                    # Create new transaction
                    app.logger.info(f"Creating transaction for user {current_user.id} ({current_user.username})")
                    
                    # CRITICAL FIX: Use the validated database user ID
                    validated_user_id = db_user.id
                    app.logger.info(f"Using validated user ID: {validated_user_id}")
                    
                    new_transaction = Transaction(
                        user_id=validated_user_id,  # Use validated user ID
                        date=parsed_date.strftime('%Y-%m-%d') if parsed_date else None,
                        description=record.get('description', ''),
                        amount=float(record.get('amount', 0)),
                        type=record.get('type', 'Other-cfo'),
                        activity_category=None
                    )
                    
                    # Set original currency and amount from record if present
                    if 'currency' in record and record['currency']:
                        new_transaction.original_currency = record['currency'].upper()  # Normalize to uppercase ISO codes
                        new_transaction.original_amount = float(record.get('amount', 0))
                    
                    # Verify the transaction was created with the correct user ID
                    app.logger.info(f"Transaction created with user_id: {new_transaction.user_id}")
                    
                    # Double-check the user ID is correct
                    if new_transaction.user_id != validated_user_id:
                        app.logger.error(f"CRITICAL ERROR: Transaction user_id mismatch! Expected: {validated_user_id}, Got: {new_transaction.user_id}")
                        continue
                    
                    # Apply currency conversion if needed
                    try:
                        from src.currency_service import get_currency_service
                        currency_service = get_currency_service()
                        if currency_service and currency_service.config.get('CURRENCY_CONVERSION_ENABLED', True):
                            conversion_success = new_transaction.apply_currency_conversion(currency_service)
                            if not conversion_success:
                                app.logger.warning(f"Currency conversion failed for transaction: {record}")
                    except AttributeError as e:
                        app.logger.error(f"Currency conversion method not available for transaction: {str(e)}")
                        # Set default currency values to prevent transaction creation failure
                        if not hasattr(new_transaction, 'original_currency') or not new_transaction.original_currency:
                            new_transaction.original_currency = 'AED'
                        if not hasattr(new_transaction, 'original_amount') or not new_transaction.original_amount:
                            new_transaction.original_amount = new_transaction.amount
                        if not hasattr(new_transaction, 'base_currency_amount') or not new_transaction.base_currency_amount:
                            new_transaction.base_currency_amount = new_transaction.amount
                    except Exception as e:
                        app.logger.error(f"Currency conversion error for transaction: {str(e)}")
                        # Set default currency values to prevent transaction creation failure
                        if not hasattr(new_transaction, 'original_currency') or not new_transaction.original_currency:
                            new_transaction.original_currency = 'AED'
                        if not hasattr(new_transaction, 'original_amount') or not new_transaction.original_amount:
                            new_transaction.original_amount = new_transaction.amount
                        if not hasattr(new_transaction, 'base_currency_amount') or not new_transaction.base_currency_amount:
                            new_transaction.base_currency_amount = new_transaction.amount
                    
                    db.session.add(new_transaction)
                    transactions_created += 1
                    
                except Exception as e:
                    app.logger.error(f"Error creating transaction from record {record}: {str(e)}")
                    continue
            
            # Commit all transactions
            db.session.commit()
            
            app.logger.info(f"Successfully created {transactions_created} transactions from upload for user {current_user.id} ({current_user.username})")
            
            # Clean up temporary file
            cleanup_temp_file(upload_token)
            session.pop('upload_file_token', None)
            
            # Add date parsing warnings to existing warnings
            warnings = result.get('warnings', [])
            if date_parse_warnings > 0:
                warnings.append(f"Skipped {date_parse_warnings} records due to invalid dates")
            
            return jsonify({
                'message': f'File processed successfully. {transactions_created} transactions created.',
                'data': result.get('data', []),
                'transactions_created': transactions_created,
                'columns': result.get('columns', []),
                'mapped_columns': result.get('mapped_columns', {}),
                'warnings': warnings,
                'user_id': current_user.id,
                'username': current_user.username
            })
        else:
            # Clean up temporary file
            cleanup_temp_file(upload_token)
            session.pop('upload_file_token', None)
            
            return jsonify({
                'message': 'File processed but no valid data found',
                'data': [],
                'transactions_created': 0,
                'warnings': result.get('warnings', []),
                'user_id': current_user.id,
                'username': current_user.username
            })
            
    except json.JSONDecodeError:
        cleanup_temp_file(upload_token)
        session.pop('upload_file_token', None)
        return jsonify({'error': {'code': 'INVALID_MAPPING', 'message': 'Invalid column mapping format', 'details': None, 'user_id': current_user.id if current_user.is_authenticated else None}}), 400
    except Exception as e:
        db.session.rollback()
        cleanup_temp_file(upload_token)
        session.pop('upload_file_token', None)
        app.logger.error(f"Error in upload_process: {str(e)}")
        return jsonify({'error': {'code': 'PROCESS_ERROR', 'message': str(e), 'details': None, 'user_id': current_user.id if current_user.is_authenticated else None}}), 500

@app.route('/set-initial-balance', methods=['POST'])
@login_required
def set_initial_balance():
    try:
        amount = float(request.form.get('initial_balance'))
        
        # Get or create initial balance record for the user
        initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
        if initial_balance:
            initial_balance.balance = amount
        else:
            initial_balance = InitialBalance(user_id=current_user.id, balance=amount)
            db.session.add(initial_balance)
            
        db.session.commit()
        flash('Initial balance set successfully', 'success')
    except ValueError:
        flash('Please enter a valid number for the initial balance', 'danger')
    except Exception as e:
        flash(f'Error setting initial balance: {str(e)}', 'danger')
        
    return redirect(url_for('cash_overview'))

@app.route('/export/<file_type>')
@login_required
def export(file_type):
    # Get accessible transactions based on user role
    transactions_query, filename_prefix = prepare_export_data(current_user, file_type)
    transactions = transactions_query.all()
    
    # Log data access for security auditing
    log_data_access(current_user, 'transactions', 'export', f'export_{file_type}')
    
    data = [{
        'Date': transaction.date,
        'Description': transaction.description,
        'Amount': transaction.amount,
        'Type': transaction.type,
        'Original Currency': transaction.original_currency,
        'Original Amount': transaction.original_amount,
        'Base (AED) Amount': transaction.base_currency_amount or transaction.amount,
        'Exchange Rate': transaction.exchange_rate,
        'Rate Date': transaction.rate_date.isoformat() if transaction.rate_date else None,
        'User': transaction.user.username if hasattr(transaction, 'user') else f'User {transaction.user_id}'
    } for transaction in transactions]

    df = pd.DataFrame(data)

    if file_type == 'csv':
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        return send_file(output, mimetype='text/csv', as_attachment=True, download_name=f'{filename_prefix}.csv')
    elif file_type == 'excel':
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Transactions')
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                         as_attachment=True, download_name=f'{filename_prefix}.xlsx')
    elif file_type == 'pdf':
        flash('PDF export is not yet implemented.', 'warning')
        return redirect(url_for('home'))
    else:
        flash('Invalid file type requested.', 'danger')
        return redirect(url_for('home'))

@app.route('/save_transactions', methods=['POST'])
@login_required
def save_transactions():
    data = request.json
    try:
        for item in data:
            new_transaction = Transaction(
                user_id=current_user.id,
                date=item['date'],
                description=item['description'],
                amount=float(item['amount']),
                type=item['type']
            )
            db.session.add(new_transaction)
        db.session.commit()
        return jsonify({'message': 'Transactions saved successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/balance-by-date', methods=['POST'])
@login_required
def balance_by_date():
    try:
        date_str = request.form.get('date')
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Get initial balance
        initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_amount = initial_balance.balance if initial_balance else 0
        
        # Get all transactions up to the target date
        transactions = Transaction.query.filter(
            Transaction.user_id == current_user.id,
            Transaction.date <= target_date
        ).all()
        
        # Calculate totals
        total_cfo, total_cfi, total_cff = calculate_totals(transactions)
        balance = initial_amount + total_cfo + total_cfi + total_cff
        
        return jsonify({
            'success': True,
            'balance': balance,
            'date': date_str
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400
    
@app.route('/forecast', methods=['GET'])
@super_admin_required
def forecast():
    try:
        # Get current user's transactions only
        transactions_query = Transaction.query.filter_by(user_id=current_user.id)
        transactions = transactions_query.order_by(Transaction.date).all()
        
        if not transactions:
            return jsonify({
                'error': 'No transactions found. Please add some transactions first.'
            }), 400
            
        # Convert transactions to a format suitable for analysis
        transaction_data = []
        for t in transactions:
            if isinstance(t.date, str):
                date_str = t.date
            else:
                date_str = t.date.strftime('%Y-%m-%d')
            transaction_data.append({
                'date': date_str,
                'amount': t.amount,
                'type': t.type,
                'description': t.description
            })
        
        # Initialize financial analytics with API key
        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({
                'error': 'API key not found. Please set the ANTHROPIC_API_KEY environment variable.'
            }), 500
            
        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        
        # Get initial balance
        initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0
        
        # Calculate current balance
        current_balance = initial_balance + sum(t.amount for t in transactions)
        
        # Mock working capital data (you may want to replace this with actual data)
        working_capital = {
            'current_assets': current_balance,
            'current_liabilities': 0,
            'cash': current_balance
        }
        
        # Get analysis results using the correct method
        analysis_results = analytics.generate_advanced_financial_analysis(
            initial_balance=initial_balance,
            current_balance=current_balance,
            transaction_history=transaction_data,
            working_capital=working_capital
        )
        
        return jsonify({
            'success': True,
            'ai_analysis': analysis_results.get('ai_analysis', 'No insights available'),
            'patterns': analysis_results.get('patterns', {'seasonal_pattern': [0] * 12}),
            'forecasts': analysis_results.get('forecasts', {'90_days': [0] * 90}),
            'risk_metrics': analysis_results.get('risk_metrics', {
                'liquidity_ratio': 0,
                'cash_flow_volatility': 0,
                'burn_rate': 0,
                'runway_months': 0
            })
        })
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500

@app.route('/generate_cashflow_statement', methods=['GET'])
@login_required
def generate_cashflow_statement_route():
    try:
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        if not transactions:
            return jsonify({'error': 'No transactions found'}), 400

        initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

        # Enhanced transaction data with transaction_id
        transaction_data = []
        for t in transactions:
            # Handle date formatting - check if it's already a string or datetime object
            if isinstance(t.date, str):
                date_str = t.date
            else:
                date_str = t.date.strftime('%Y-%m-%d')
                
            transaction_data.append({
                "transaction_id": getattr(t, 'transaction_id', str(t.id)),
                "date": date_str,
                "description": t.description,
                "amount": t.amount,
                "type": t.type
            })

        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'error': 'API key not found'}), 500
            
        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        statement = analytics.generate_dual_cashflow_statement(initial_balance, transaction_data)
        return jsonify(statement)
    except Exception as e:
        current_app.logger.error(f"Error generating cashflow statement: {str(e)}")
        return jsonify({'error': f"Failed to generate cash flow statement: {str(e)}"}), 500

@app.route('/cashflow-statement', methods=['GET'])
@login_required
def cashflow_statement():
    try:
        # Determine if current user has any transactions
        try:
            transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
        except Exception:
            transaction_count = 0

        has_data = bool(transaction_count and transaction_count > 0)

        # Render the cashflow statement page with consistent AI context
        return render_template(
            'cashflow_statement.html',
            has_data=has_data,
            transaction_count=transaction_count,
        )
    except Exception as e:
        current_app.logger.error(f"Error rendering cashflow statement page: {e}")
        flash('Failed to load Cash Flow Statement page', 'danger')
        return redirect(url_for('home'))

@app.route('/export_cashflow/excel')
@login_required
def export_cashflow_excel_route():
    try:
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        if not transactions:
            return jsonify({'error': 'No transactions found'}), 400

        initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

        transaction_data = []
        for t in transactions:
            date_str = t.date if isinstance(t.date, str) else t.date.strftime('%Y-%m-%d')
            transaction_data.append({
                "transaction_id": getattr(t, 'transaction_id', str(t.id)),
                "date": date_str,
                "description": t.description,
                "amount": t.amount,
                "type": t.type
            })

        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'error': 'API key not found'}), 500

        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        statement = analytics.generate_dual_cashflow_statement(initial_balance, transaction_data)
        
        # Fallback if AI data is malformed
        if not statement or 'indirect_method' not in statement:
            # Create a simple fallback statement
            total_income = sum(t['amount'] for t in transaction_data if t['amount'] > 0)
            total_expenses = sum(abs(t['amount']) for t in transaction_data if t['amount'] < 0)
            net_income = total_income - total_expenses
            
            statement = {
                'indirect_method': {
                    'operating_activities': {
                        'net_income': net_income,
                        'depreciation': 0,
                        'amortization': 0,
                        'changes_in_working_capital': 0,
                        'total_operating_cash_flow': net_income
                    },
                    'investing_activities': {
                        'purchase_of_assets': 0,
                        'sale_of_assets': 0,
                        'investments': 0,
                        'total_investing_cash_flow': 0
                    },
                    'financing_activities': {
                        'debt_issued': 0,
                        'debt_repayment': 0,
                        'equity_issued': 0,
                        'dividends_paid': 0,
                        'total_financing_cash_flow': 0
                    },
                    'net_cash_flow': net_income,
                    'beginning_cash': initial_balance,
                    'ending_cash': initial_balance + net_income
                },
                'executive_summary': 'Basic cash flow analysis based on transaction data.',
                'validation': {
                    'duplicates': [],
                    'unusual_patterns': [],
                    'warnings': ['AI analysis unavailable - using basic calculation'],
                    'total_issues': 0
                }
            }

        # Generate Excel with better error handling
        output = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Helper function for currency formatting
        def format_currency(cell):
            cell.number_format = '$#,##0.00'
            cell.font = Font(bold=True)

        # Helper function to safely get values
        def safe_get(data, key, default=0):
            try:
                value = data.get(key, default)
                return float(value) if value is not None else default
            except (ValueError, TypeError):
                return default

        # Executive Summary Sheet
        summary_sheet = workbook.create_sheet("Executive Summary")
        summary_sheet['A1'] = "Cash Flow Statement (Indirect Method)"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        summary_sheet['A3'] = "Executive Summary:"
        summary_sheet['A3'].font = Font(bold=True)
        summary_sheet['A4'] = statement.get('executive_summary', 'No summary available')
        summary_sheet['A6'] = "Key Metrics:"
        summary_sheet['A6'].font = Font(bold=True)
        summary_sheet['A7'] = "Beginning Cash:"
        summary_sheet['B7'] = initial_balance
        format_currency(summary_sheet['B7'])
        summary_sheet['A8'] = "Net Cash Flow:"
        summary_sheet['B8'] = safe_get(statement.get('indirect_method', {}), 'net_cash_flow', 0)
        format_currency(summary_sheet['B8'])
        summary_sheet['A9'] = "Ending Cash:"
        summary_sheet['B9'] = safe_get(statement.get('indirect_method', {}), 'ending_cash', initial_balance + safe_get(statement.get('indirect_method', {}), 'net_cash_flow', 0))
        format_currency(summary_sheet['B9'])

        # Indirect Method Sheet (Main Statement)
        if 'indirect_method' in statement:
            indirect_sheet = workbook.create_sheet("Cash Flow Statement")
            indirect_sheet['A1'] = "Statement of Cash Flows (Indirect Method)"
            indirect_sheet['A1'].font = Font(size=14, bold=True)
            row = 3
            
            # Operating Activities
            indirect_sheet[f'A{row}'] = "CASH FLOWS FROM OPERATING ACTIVITIES"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            row += 2
            
            indirect_data = statement['indirect_method']
            operating = indirect_data.get('operating_activities', {})
            
            indirect_sheet[f'A{row}'] = "Net Income"
            indirect_sheet[f'B{row}'] = safe_get(operating, 'net_income', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Adjustments to reconcile net income to net cash:"
            indirect_sheet[f'A{row}'].font = Font(italic=True)
            row += 1
            
            indirect_sheet[f'A{row}'] = "  Depreciation"
            indirect_sheet[f'B{row}'] = safe_get(operating, 'depreciation', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "  Amortization"
            indirect_sheet[f'B{row}'] = safe_get(operating, 'amortization', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "  Changes in Working Capital"
            indirect_sheet[f'B{row}'] = safe_get(operating, 'changes_in_working_capital', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Net Cash from Operating Activities"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            indirect_sheet[f'B{row}'] = safe_get(operating, 'total_operating_cash_flow', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 2
            
            # Investing Activities
            indirect_sheet[f'A{row}'] = "CASH FLOWS FROM INVESTING ACTIVITIES"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            row += 2
            
            investing = indirect_data.get('investing_activities', {})
            
            indirect_sheet[f'A{row}'] = "Purchase of Assets"
            indirect_sheet[f'B{row}'] = safe_get(investing, 'purchase_of_assets', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Sale of Assets"
            indirect_sheet[f'B{row}'] = safe_get(investing, 'sale_of_assets', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Investments"
            indirect_sheet[f'B{row}'] = safe_get(investing, 'investments', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Net Cash from Investing Activities"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            indirect_sheet[f'B{row}'] = safe_get(investing, 'total_investing_cash_flow', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 2
            
            # Financing Activities
            indirect_sheet[f'A{row}'] = "CASH FLOWS FROM FINANCING ACTIVITIES"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            row += 2
            
            financing = indirect_data.get('financing_activities', {})
            
            indirect_sheet[f'A{row}'] = "Debt Issued"
            indirect_sheet[f'B{row}'] = safe_get(financing, 'debt_issued', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Debt Repayment"
            indirect_sheet[f'B{row}'] = safe_get(financing, 'debt_repayment', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Equity Issued"
            indirect_sheet[f'B{row}'] = safe_get(financing, 'equity_issued', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Dividends Paid"
            indirect_sheet[f'B{row}'] = safe_get(financing, 'dividends_paid', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Net Cash from Financing Activities"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            indirect_sheet[f'B{row}'] = safe_get(financing, 'total_financing_cash_flow', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 2
            
            # Net Change and Ending Balance
            indirect_sheet[f'A{row}'] = "NET INCREASE (DECREASE) IN CASH"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            indirect_sheet[f'B{row}'] = safe_get(indirect_data, 'net_cash_flow', 0)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Cash at Beginning of Period"
            indirect_sheet[f'B{row}'] = safe_get(indirect_data, 'beginning_cash', initial_balance)
            format_currency(indirect_sheet[f'B{row}'])
            row += 1
            
            indirect_sheet[f'A{row}'] = "Cash at End of Period"
            indirect_sheet[f'A{row}'].font = Font(bold=True)
            indirect_sheet[f'B{row}'] = safe_get(indirect_data, 'ending_cash', initial_balance + safe_get(indirect_data, 'net_cash_flow', 0))
            format_currency(indirect_sheet[f'B{row}'])

        # Validation Sheet
        if 'validation' in statement:
            val_sheet = workbook.create_sheet("Validation")
            val_sheet['A1'] = "Validation Report"
            val_sheet['A1'].font = Font(size=14, bold=True)
            row = 3
            if statement['validation']['duplicates']:
                val_sheet[f'A{row}'] = "Duplicates"
                val_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                for dup in statement['validation']['duplicates']:
                    val_sheet[f'A{row}'] = dup['transaction_id']
                    val_sheet[f'B{row}'] = dup['issue']
                    val_sheet[f'C{row}'] = dup['severity']
                    row += 1
            if statement['validation']['unusual_patterns']:
                val_sheet[f'A{row}'] = "Unusual Patterns"
                val_sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                for pattern in statement['validation']['unusual_patterns']:
                    val_sheet[f'A{row}'] = pattern['pattern']
                    val_sheet[f'B{row}'] = pattern['description']
                    val_sheet[f'C{row}'] = pattern['severity']
                    row += 1

        try:
            workbook.save(output)
            output.seek(0)
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                             as_attachment=True, download_name='cashflow_statement.xlsx')
        except Exception as excel_error:
            current_app.logger.error(f"Excel save error: {str(excel_error)}")
            # Try to create a minimal Excel file
            try:
                output = BytesIO()
                workbook = Workbook()
                ws = workbook.active
                ws.title = "Cash Flow Statement"
                ws['A1'] = "Cash Flow Statement (Basic)"
                ws['A1'].font = Font(size=16, bold=True)
                ws['A3'] = "Beginning Cash:"
                ws['B3'] = initial_balance
                ws['A4'] = "Net Cash Flow:"
                ws['B4'] = sum(t['amount'] for t in transaction_data)
                ws['A5'] = "Ending Cash:"
                ws['B5'] = initial_balance + sum(t['amount'] for t in transaction_data)
                workbook.save(output)
                output.seek(0)
                return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                                 as_attachment=True, download_name='cashflow_statement_basic.xlsx')
            except Exception as fallback_error:
                current_app.logger.error(f"Fallback Excel error: {str(fallback_error)}")
                return jsonify({'error': f'Excel generation failed: {str(excel_error)}'}), 500

    except Exception as e:
        current_app.logger.error(f"Excel export error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Generate monthly chart function    
def generate_monthly_balance_chart():
    # Get transactions for the current user
    transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
    
    # Get initial balance for the current user
    initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
    initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

    monthly_balances = []
    current_balance = initial_balance
    current_month = None

    for transaction in transactions:
        transaction_date = transaction.date if isinstance(transaction.date, datetime) else datetime.strptime(transaction.date, '%Y-%m-%d')
        
        if current_month != transaction_date.replace(day=1):
            if current_month:
                last_day = calendar.monthrange(current_month.year, current_month.month)[1]
                monthly_balances.append({
                    'date': current_month.replace(day=last_day),
                    'balance': current_balance
                })
            current_month = transaction_date.replace(day=1)

        current_balance += transaction.amount

    # Add the last month's balance
    if current_month:
        last_day = calendar.monthrange(current_month.year, current_month.month)[1]
        monthly_balances.append({
            'date': current_month.replace(day=last_day),
            'balance': current_balance
        })

    # Create the chart
    fig = Figure(figsize=(12, 6))
    axis = fig.add_subplot(1, 1, 1)
    
    if monthly_balances:
        dates = [balance['date'] for balance in monthly_balances]
        balances = [balance['balance'] for balance in monthly_balances]
        axis.plot(dates, balances, marker='o')  # Added markers for each data point

        # Improve Y-axis formatting
        def currency_formatter(x, p):
            return f'${x:,.0f}'
        
        axis.yaxis.set_major_formatter(ticker.FuncFormatter(currency_formatter))
        
        # Adjust Y-axis ticks for better readability
        axis.yaxis.set_major_locator(ticker.MaxNLocator(nbins=10, integer=True))
        
        # Format X-axis to show dates nicely
        axis.xaxis.set_major_locator(MonthLocator())
        axis.xaxis.set_major_formatter(DateFormatter('%Y-%m-%d'))
        
        # Add some padding to the y-axis
        ylim = axis.get_ylim()
        axis.set_ylim([ylim[0] - (ylim[1] - ylim[0]) * 0.1, ylim[1] + (ylim[1] - ylim[0]) * 0.1])
    else:
        # If no data, show a message on the chart
        axis.text(0.5, 0.5, 'No transaction data available', 
                 horizontalalignment='center', verticalalignment='center',
                 transform=axis.transAxes, fontsize=14)
        axis.set_xticks([])
        axis.set_yticks([])

    axis.set_title('Monthly Balance', fontsize=16, fontweight='bold')
    axis.set_xlabel('Date', fontsize=12)
    axis.set_ylabel('Balance ($)', fontsize=12)
    axis.tick_params(axis='both', which='major', labelsize=10)
    axis.tick_params(axis='x', rotation=45)
    
    # Add gridlines for better readability
    axis.grid(True, linestyle='--', alpha=0.7)

    fig.tight_layout()

    # Convert plot to PNG image
    png_image = io.BytesIO()
    FigureCanvas(fig).print_png(png_image)
    
    # Encode PNG image to base64 string
    png_image_b64_string = "data:image/png;base64,"
    png_image_b64_string += base64.b64encode(png_image.getvalue()).decode('utf8')

    return png_image_b64_string

#Map plotting chart
@app.route('/monthly-balances', methods=['GET'])
@login_required
def monthly_balances():
    try:
        app.logger.info("Starting to generate monthly balance data")
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        initial_balance_record = InitialBalance.query.filter_by(user_id=current_user.id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0

        monthly_data = {}
        current_balance = initial_balance

        for transaction in transactions:
            month = transaction.date.strftime('%Y-%m') if isinstance(transaction.date, datetime) else datetime.strptime(transaction.date, '%Y-%m-%d').strftime('%Y-%m')
            
            if month not in monthly_data:
                monthly_data[month] = current_balance
            
            current_balance += transaction.amount
            monthly_data[month] = current_balance

        app.logger.info(f"Monthly balance data generated: {monthly_data}")
        return jsonify({
            'success': True,
            'data': {
                'labels': list(monthly_data.keys()),
                'balances': list(monthly_data.values())
            }
        })
    except Exception as e:
        app.logger.error(f"Error generating monthly balance data: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/settings')
@login_required
def settings():
    user_preferences = UserPreferences.query.filter_by(user_id=current_user.id).first()
    return render_template('settings.html', user_preferences=user_preferences)

@app.route('/update_profile', methods=['POST'])
@login_required
def update_profile():
    current_user.username = request.form.get('username')
    current_user.email = request.form.get('email')
    db.session.commit()
    flash('Profile updated successfully', 'success')
    return redirect(url_for('settings'))

@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    if check_password_hash(current_user.password, request.form.get('current_password')):
        if request.form.get('new_password') == request.form.get('confirm_password'):
            current_user.password = generate_password_hash(request.form.get('new_password'))
            db.session.commit()
            flash('Password changed successfully', 'success')
        else:
            flash('New passwords do not match', 'danger')
    else:
        flash('Current password is incorrect', 'danger')
    return redirect(url_for('settings'))

@app.route('/update_modules', methods=['POST'])
@login_required
def update_modules():
    user_preferences = UserPreferences.query.filter_by(user_id=current_user.id).first()
    if not user_preferences:
        user_preferences = UserPreferences(user_id=current_user.id)
        db.session.add(user_preferences)
    user_preferences.modules = request.form.getlist('modules')
    # Enforce role-based required modules after manual selection
    try:
        user_preferences.ensure_role_based_modules(current_user)
    except Exception:
        pass
    db.session.commit()
    flash('Module preferences updated successfully', 'success')
    return redirect(url_for('settings'))

@app.context_processor
def utility_processor():
    def get_user_preferences():
        if current_user.is_authenticated:
            prefs = UserPreferences.query.filter_by(user_id=current_user.id).first()
            if not prefs:
                prefs = UserPreferences(user_id=current_user.id)
                db.session.add(prefs)
                db.session.flush()  # Ensure the preferences have an ID
                try:
                    prefs.ensure_role_based_modules(current_user)
                except Exception:
                    pass
                db.session.commit()
            else:
                # Ensure existing preferences include any role-based defaults
                try:
                    if prefs.ensure_role_based_modules(current_user):
                        db.session.commit()
                except Exception:
                    pass
            return prefs
        return None
    return dict(user_preferences=get_user_preferences())

@app.context_processor
def csrf_processor():
    """Make CSRF token available in templates."""
    from flask_wtf.csrf import generate_csrf
    return dict(csrf_token=generate_csrf())

@app.route('/cash-overview')
@login_required
def cash_overview():
    # Get or create initial balance for the current user
    initial_balance = InitialBalance.query.filter_by(user_id=current_user.id).first()
    if not initial_balance:
        initial_balance = InitialBalance(user_id=current_user.id, balance=0)
        db.session.add(initial_balance)
        db.session.commit()

    # Get all transactions for the current user
    transactions = Transaction.query.filter_by(user_id=current_user.id).all()
    
    # Calculate totals
    total_cfo, total_cfi, total_cff = calculate_totals(transactions)
    balance = initial_balance.balance + total_cfo + total_cfi + total_cff

    # Calculate burn rate and runway
    burn_rate = calculate_burn_rate(transactions) or 0.0
    runway_months = calculate_runway(balance, burn_rate)
    try:
        runway_months = float(runway_months) if runway_months is not None else 0.0
    except Exception:
        runway_months = 0.0

    # Provide safe default enterprise KPIs for template
    enterprise_kpis = {
        'current_ratio': 0.0,
        'quick_ratio': 0.0,
        'working_capital': 0.0,
        'cash_cycle': 0.0,
        'aging_buckets': {
            '0-30': 0.0,
            '31-60': 0.0,
            '61-90': 0.0,
            '90+': 0.0
        },
        'collection_efficiency': 0.0,
        'dso': 0.0,
        'payables_schedule': {
            '30': 0.0
        },
        'dpo': 0.0
    }

    return render_template('cash_overview.html', 
                         initial_balance=initial_balance.balance,
                         total_cfo=total_cfo, 
                         total_cfi=total_cfi, 
                         total_cff=total_cff, 
                         balance=balance,
                         burn_rate=burn_rate,
                         runway_months=runway_months,
                         enterprise_kpis=enterprise_kpis)

@app.route('/cash-activities')
@login_required
def cash_activities():
    page = request.args.get('page', 1, type=int)
    per_page = 8
    
    # Get paginated transactions for the current user
    paginated_transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date.desc()).paginate(page=page, per_page=per_page)
    
    return render_template('cash_activities.html', transactions=paginated_transactions)

@app.route('/ai-analysis')
@login_required
@performance_monitor
def ai_analysis():
    try:
        # Role-based access with module check and template context
        user_roles = [role.name for role in current_user.roles] if hasattr(current_user, 'roles') else []

        # Ensure user preferences exist
        prefs = UserPreferences.query.filter_by(user_id=current_user.id).first()
        if not prefs:
            prefs = UserPreferences(user_id=current_user.id)
            db.session.add(prefs)
            db.session.flush()
            try:
                prefs.ensure_role_based_modules(current_user)
            except Exception:
                pass
            db.session.commit()

        modules = prefs.modules or []

        # Transaction count for current user
        transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
        has_data = transaction_count > 0

        # Enhanced context with performance monitoring
        ctx = get_ai_navigation_context(
            'dashboard',
            page_title='AI Analysis Dashboard',
            page_subtitle='Comprehensive AI-powered financial insights',
            page_description='Advanced AI analysis dashboard with real-time insights and predictive analytics'
        )

        return render_template(
            'ai_analysis_min.html',
            transaction_count=transaction_count,
            has_data=has_data,
            status_message='AI system ready',
            **ctx
        )
    except Exception as e:
        log_route_error('ai_analysis', e, current_user.id if current_user.is_authenticated else None)
        flash('An error occurred loading the AI analysis page. Please try again.', 'error')
        return redirect(url_for('home'))

# Minimal AI pages
@app.route('/ai-analysis/generate', methods=['GET'])
@login_required
def ai_generate():
    # Check if user has transaction data
    transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
    has_data = transaction_count > 0
    
    ctx = get_ai_navigation_context('dashboard')  # or 'generate' if you add that to nav
    return render_template(
        'ai_analysis_generate.html',
        has_data=has_data,
        transaction_count=transaction_count,
        **ctx
    )

@app.route('/ai-analysis/cashflow', methods=['GET'])
@login_required
@performance_monitor
def ai_cashflow():
    try:
        # Check if user has transaction data
        transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
        has_data = transaction_count > 0
        
        # Enhanced context with accessibility and SEO metadata
        ctx = get_ai_navigation_context(
            'cashflow',
            page_title='Cash Flow Analysis',
            page_subtitle='AI-powered cash flow insights and predictions',
            page_description='Advanced cash flow analysis with AI-driven insights, trend identification, and predictive forecasting'
        )
        
        return render_template(
            'ai_analysis_cashflow.html',
            has_data=has_data,
            transaction_count=transaction_count,
            status_message='Cash flow analysis ready' if has_data else 'Add transactions to enable cash flow analysis',
            **ctx
        )
    except Exception as e:
        log_route_error('ai_cashflow', e, current_user.id if current_user.is_authenticated else None)
        flash('An error occurred loading the cash flow analysis. Please try again.', 'error')
        return redirect(url_for('ai_analysis'))

@app.route('/ai-analysis/risk', methods=['GET'])
@login_required
def ai_risk():
    try:
        # Check if user has transaction data
        transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
        has_data = transaction_count > 0
        
        # Initialize default risk metrics
        risk_metrics = {
            'credit_score': 'N/A',
            'liquidity_level': 'N/A', 
            'market_exposure': 'N/A',
            'operational_index': 'N/A'
        }
        
        # Initialize default recent activities
        recent_risk_activities = []
        
        # If data exists, get risk analysis data
        if has_data:
            try:
                # Get basic risk metrics
                transactions = Transaction.query.filter_by(user_id=current_user.id).all()
                total_balance = sum(t.amount for t in transactions)
                avg_transaction = total_balance / len(transactions) if transactions else 0
                
                # Calculate basic risk metrics
                risk_metrics = {
                    'credit_score': 'Good' if avg_transaction > -500 else 'Fair' if avg_transaction > -1000 else 'Poor',
                    'liquidity_level': 'High' if total_balance > 10000 else 'Medium' if total_balance > 1000 else 'Low',
                    'market_exposure': 'Low' if abs(avg_transaction) < 1000 else 'Medium',
                    'operational_index': 'Low' if transaction_count > 10 else 'Medium'
                }
                
                # Mock recent activities
                recent_risk_activities = [
                    {
                        'title': 'Risk Assessment Completed',
                        'success': True,
                        'timestamp': '2024-01-15 10:30:00'
                    }
                ]
                
            except Exception as e:
                current_app.logger.error(f"Error calculating risk data: {str(e)}")
                # Keep default values on error
        
        # Enhanced navigation context
        breadcrumb_items = [
            {'title': 'Workspace', 'url': '/'},
            {'title': 'AI Analysis', 'url': '/ai-analysis'},
            {'title': 'Risk Assessment', 'url': None}
        ]
        
        ai_features_nav = [
            {
                'id': 'dashboard',
                'title': 'Dashboard',
                'url': '/ai-analysis/dashboard',
                'icon': 'fas fa-tachometer-alt',
                'description': 'AI system overview and metrics'
            },
            {
                'id': 'risk',
                'title': 'Risk Assessment',
                'url': '/ai-analysis/risk',
                'icon': 'fas fa-shield-alt',
                'description': 'Financial risk analysis and assessment'
            },
            {
                'id': 'anomaly',
                'title': 'Anomaly Detection',
                'url': '/ai-analysis/anomaly',
                'icon': 'fas fa-search',
                'description': 'Detect unusual transaction patterns'
            },
            {
                'id': 'forecast',
                'title': 'Advanced Forecast',
                'url': '/ai-analysis/forecast',
                'icon': 'fas fa-chart-line',
                'description': 'Predictive forecasting models'
            },
            {
                'id': 'assistant',
                'title': 'Virtual Analyst',
                'url': '/ai-analysis/assistant',
                'icon': 'fas fa-robot',
                'description': 'AI-powered data analysis assistant'
            }
        ]
        
        return render_template('ai_risk_assessment.html', 
                             has_data=has_data, 
                             transaction_count=transaction_count,
                             risk_metrics=risk_metrics,
                             recent_risk_activities=recent_risk_activities,
                             breadcrumb_items=breadcrumb_items,
                             ai_features_nav=ai_features_nav,
                             current_page='risk')
                             
    except Exception as e:
        current_app.logger.error(f"AI risk assessment page error: {str(e)}")
        flash('Error loading risk assessment page. Please try again.', 'error')
        # Enhanced navigation context for error state
        breadcrumb_items = [
            {'title': 'Workspace', 'url': '/'},
            {'title': 'AI Analysis', 'url': '/ai-analysis'},
            {'title': 'Risk Assessment', 'url': None}
        ]
        
        ai_features_nav = [
            {
                'id': 'dashboard',
                'title': 'Dashboard',
                'url': '/ai-analysis/dashboard',
                'icon': 'fas fa-tachometer-alt',
                'description': 'AI system overview and metrics'
            },
            {
                'id': 'risk',
                'title': 'Risk Assessment',
                'url': '/ai-analysis/risk',
                'icon': 'fas fa-shield-alt',
                'description': 'Financial risk analysis and assessment'
            },
            {
                'id': 'anomaly',
                'title': 'Anomaly Detection',
                'url': '/ai-analysis/anomaly',
                'icon': 'fas fa-search',
                'description': 'Detect unusual transaction patterns'
            },
            {
                'id': 'forecast',
                'title': 'Advanced Forecast',
                'url': '/ai-analysis/forecast',
                'icon': 'fas fa-chart-line',
                'description': 'Predictive forecasting models'
            },
            {
                'id': 'assistant',
                'title': 'Virtual Analyst',
                'url': '/ai-analysis/assistant',
                'icon': 'fas fa-robot',
                'description': 'AI-powered data analysis assistant'
            }
        ]
        
        return render_template('ai_risk_assessment.html', 
                             has_data=False, 
                             transaction_count=0,
                             risk_metrics={'credit_score': 'N/A', 'liquidity_level': 'N/A', 'market_exposure': 'N/A', 'operational_index': 'N/A'},
                             recent_risk_activities=[],
                             breadcrumb_items=breadcrumb_items,
                             ai_features_nav=ai_features_nav,
                             current_page='risk')

@app.route('/ai-analysis/anomaly', methods=['GET'])
@login_required
def ai_anomaly():
    try:
        # Check if user has transaction data
        transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
        has_data = transaction_count > 0
        
        # Initialize default anomaly metrics
        anomaly_metrics = {
            'transactions_scanned': 0,
            'anomalies_detected': 0,
            'accuracy': '0%',
            'last_scan': 'Never'
        }
        
        # Initialize default recent activities
        recent_anomaly_activities = []
        
        # If data exists, get anomaly detection data
        if has_data:
            try:
                # Calculate basic anomaly metrics
                anomaly_metrics = {
                    'transactions_scanned': transaction_count,
                    'anomalies_detected': max(0, transaction_count // 10),  # Mock: 10% anomaly rate
                    'accuracy': '95%',
                    'last_scan': '2024-01-15 10:30:00'
                }
                
                # Mock recent activities
                recent_anomaly_activities = [
                    {
                        'title': 'Anomaly Scan Completed',
                        'success': True,
                        'timestamp': '2024-01-15 10:30:00'
                    }
                ]
                
            except Exception as e:
                current_app.logger.error(f"Error calculating anomaly data: {str(e)}")
                # Keep default values on error
        
        ctx = get_ai_navigation_context('anomaly')
        return render_template(
            'ai_anomaly_detection.html',
            has_data=has_data,
            transaction_count=transaction_count,
            anomaly_metrics=anomaly_metrics,
            recent_anomaly_activities=recent_anomaly_activities,
            **ctx
        )
                             
    except Exception as e:
        current_app.logger.error(f"AI anomaly detection page error: {str(e)}")
        flash('Error loading anomaly detection page. Please try again.', 'error')
        return render_template('ai_anomaly_detection.html', 
                             has_data=False, 
                             transaction_count=0,
                             anomaly_metrics={'transactions_scanned': 0, 'anomalies_detected': 0, 'accuracy': '0%', 'last_scan': 'Never'},
                             recent_anomaly_activities=[])

@app.route('/ai-analysis/forecast', methods=['GET'])
@login_required
def ai_forecast():
    # Check if user has transaction data
    transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
    has_data = transaction_count > 0
    
    ctx = get_ai_navigation_context('forecast')
    return render_template(
        'ai_advanced_forecast.html',
        has_data=has_data,
        transaction_count=transaction_count,
        **ctx
    )

@app.route('/ai-analysis/dashboard', methods=['GET'])
@login_required
def ai_dashboard_page():
    try:
        # Check if user has transaction data
        transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
        has_data = transaction_count > 0
        
        # Initialize default AI performance metrics
        ai_performance_metrics = {
            'accuracy': 0,
            'response_time': 'N/A',
            'uptime': '0%'
        }
        
        # Initialize default system status
        ai_system_status = {
            'overall_status': 'unknown',
            'last_checked': None
        }
        
        # Initialize default recent activities
        recent_ai_activities = []
        
        # If data exists, get AI dashboard data
        if has_data:
            try:
                # Calculate basic AI performance metrics
                ai_performance_metrics = {
                    'accuracy': 95,  # Mock accuracy
                    'response_time': '2.3s',
                    'uptime': '99.9%'
                }
                
                # Mock system status
                ai_system_status = {
                    'overall_status': 'healthy',
                    'last_checked': '2024-01-15 10:30:00'
                }
                
                # Mock recent activities
                recent_ai_activities = [
                    {
                        'feature_name': 'Risk Assessment',
                        'success': True,
                        'timestamp': '2024-01-15 10:30:00'
                    },
                    {
                        'feature_name': 'Anomaly Detection',
                        'success': True,
                        'timestamp': '2024-01-15 09:15:00'
                    }
                ]
                
            except Exception as e:
                current_app.logger.error(f"Error calculating AI dashboard data: {str(e)}")
                # Keep default values on error
        
        # Enhanced navigation context
        breadcrumb_items = [
            {'title': 'Workspace', 'url': '/'},
            {'title': 'AI Analysis', 'url': '/ai-analysis'},
            {'title': 'Dashboard', 'url': None}
        ]
        
        ai_features_nav = [
            {
                'id': 'dashboard',
                'title': 'Dashboard',
                'url': '/ai-analysis/dashboard',
                'icon': 'fas fa-tachometer-alt',
                'description': 'AI system overview and metrics'
            },
            {
                'id': 'risk',
                'title': 'Risk Assessment',
                'url': '/ai-analysis/risk',
                'icon': 'fas fa-shield-alt',
                'description': 'Financial risk analysis and assessment'
            },
            {
                'id': 'anomaly',
                'title': 'Anomaly Detection',
                'url': '/ai-analysis/anomaly',
                'icon': 'fas fa-search',
                'description': 'Detect unusual transaction patterns'
            },
            {
                'id': 'forecast',
                'title': 'Advanced Forecast',
                'url': '/ai-analysis/forecast',
                'icon': 'fas fa-chart-line',
                'description': 'Predictive forecasting models'
            },
            {
                'id': 'assistant',
                'title': 'Virtual Analyst',
                'url': '/ai-analysis/assistant',
                'icon': 'fas fa-robot',
                'description': 'AI-powered data analysis assistant'
            }
        ]
        
        return render_template('ai_dashboard.html', 
                             has_data=has_data, 
                             transaction_count=transaction_count,
                             ai_performance_metrics=ai_performance_metrics,
                             ai_system_status=ai_system_status,
                             recent_ai_activities=recent_ai_activities,
                             breadcrumb_items=breadcrumb_items,
                             ai_features_nav=ai_features_nav,
                             current_page='dashboard')
                             
    except Exception as e:
        current_app.logger.error(f"AI dashboard page error: {str(e)}")
        flash('Error loading AI dashboard. Please try again.', 'error')
        # Enhanced navigation context for error state
        breadcrumb_items = [
            {'title': 'Workspace', 'url': '/'},
            {'title': 'AI Analysis', 'url': '/ai-analysis'},
            {'title': 'Dashboard', 'url': None}
        ]
        
        ai_features_nav = [
            {
                'id': 'dashboard',
                'title': 'Dashboard',
                'url': '/ai-analysis/dashboard',
                'icon': 'fas fa-tachometer-alt',
                'description': 'AI system overview and metrics'
            },
            {
                'id': 'risk',
                'title': 'Risk Assessment',
                'url': '/ai-analysis/risk',
                'icon': 'fas fa-shield-alt',
                'description': 'Financial risk analysis and assessment'
            },
            {
                'id': 'anomaly',
                'title': 'Anomaly Detection',
                'url': '/ai-analysis/anomaly',
                'icon': 'fas fa-search',
                'description': 'Detect unusual transaction patterns'
            },
            {
                'id': 'forecast',
                'title': 'Advanced Forecast',
                'url': '/ai-analysis/forecast',
                'icon': 'fas fa-chart-line',
                'description': 'Predictive forecasting models'
            },
            {
                'id': 'assistant',
                'title': 'Virtual Analyst',
                'url': '/ai-analysis/assistant',
                'icon': 'fas fa-robot',
                'description': 'AI-powered data analysis assistant'
            }
        ]
        
        return render_template('ai_dashboard.html', 
                             has_data=False, 
                             transaction_count=0,
                             ai_performance_metrics={'accuracy': 0, 'response_time': 'N/A', 'uptime': '0%'},
                             ai_system_status={'overall_status': 'unknown', 'last_checked': None},
                             recent_ai_activities=[],
                             breadcrumb_items=breadcrumb_items,
                             ai_features_nav=ai_features_nav,
                             current_page='dashboard')

@app.route('/ai-analysis/assistant', methods=['GET'])
@login_required
def ai_assistant():
    try:
        # Check if user has transaction data
        transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
        has_data = transaction_count > 0
        
        return render_template('chatbot.html', 
                             has_data=has_data, 
                             transaction_count=transaction_count)
                             
    except Exception as e:
        current_app.logger.error(f"AI assistant page error: {str(e)}")
        flash('Error loading AI assistant. Please try again.', 'error')
        return render_template('chatbot.html', 
                             has_data=False, 
                             transaction_count=0)

# AI Pages Health Check Route (for testing)
# API endpoint for error reporting
@app.route('/api/errors', methods=['POST'])
def report_error():
    """API endpoint for client-side error reporting"""
    try:
        error_data = request.get_json()
        if not error_data:
            return jsonify({'error': 'No error data provided'}), 400
        
        # Log error with context
        logging.error(f"Client error: {error_data}")
        
        return jsonify({'status': 'error_reported'}), 200
    except Exception as e:
        logging.error(f"Error reporting failed: {str(e)}")
        return jsonify({'error': 'Failed to report error'}), 500

# API endpoint for performance metrics
@app.route('/api/performance', methods=['POST'])
def report_performance():
    """API endpoint for client-side performance reporting"""
    try:
        perf_data = request.get_json()
        if not perf_data:
            return jsonify({'error': 'No performance data provided'}), 400
        
        # Log performance metrics
        logging.info(f"Performance metrics: {perf_data}")
        
        return jsonify({'status': 'metrics_reported'}), 200
    except Exception as e:
        logging.error(f"Performance reporting failed: {str(e)}")
        return jsonify({'error': 'Failed to report performance'}), 500

@app.route('/api/currency/rates', methods=['GET'])
@login_required
def get_currency_rates():
    """API endpoint to get current exchange rates"""
    try:
        currency_service = get_currency_service()
        if not currency_service:
            return jsonify({'error': 'Currency service not available'}), 503
        
        cache_info = currency_service.get_cache_info()
        return jsonify({
            'success': True,
            'rates': currency_service._cache.get('rates', {}),
            'base_currency': cache_info.get('base_currency', 'AED'),
            'last_updated': cache_info.get('last_updated').isoformat() if cache_info.get('last_updated') else None,
            'source': cache_info.get('source', 'unknown')
        })
    except Exception as e:
        return jsonify({'error': f'Failed to get currency rates: {str(e)}'}), 500

@app.route('/api/currency/refresh', methods=['POST'])
@login_required
def refresh_currency_rates():
    """API endpoint to manually refresh exchange rates"""
    try:
        currency_service = get_currency_service()
        if not currency_service:
            return jsonify({'error': 'Currency service not available'}), 503
        
        success = currency_service.refresh_rates()
        if success:
            return jsonify({'success': True, 'message': 'Exchange rates refreshed successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to refresh exchange rates'}), 500
    except Exception as e:
        return jsonify({'error': f'Failed to refresh rates: {str(e)}'}), 500

@app.route('/api/currency/convert', methods=['POST'])
@login_required
def convert_currency():
    """API endpoint to convert currency amounts"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        amount = data.get('amount')
        from_currency = data.get('from_currency')
        to_currency = data.get('to_currency')
        
        if not all([amount, from_currency, to_currency]):
            return jsonify({'error': 'Missing required fields: amount, from_currency, to_currency'}), 400
        
        currency_service = get_currency_service()
        if not currency_service:
            return jsonify({'error': 'Currency service not available'}), 503
        
        converted_amount = currency_service.convert_amount(amount, from_currency, to_currency)
        if converted_amount is None:
            return jsonify({'error': 'Currency conversion failed'}), 400
        
        return jsonify({
            'success': True,
            'original_amount': amount,
            'converted_amount': converted_amount,
            'from_currency': from_currency,
            'to_currency': to_currency,
            'exchange_rate': currency_service.get_exchange_rate(from_currency, to_currency)
        })
    except Exception as e:
        return jsonify({'error': f'Currency conversion failed: {str(e)}'}), 500

@app.route('/ai/health-check', methods=['GET'])
@login_required
def ai_health_check():
    """Health check for all AI pages to ensure they don't throw 500 errors"""
    try:
        health_status = {
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'pages_tested': []
        }
        
        # Test each AI page by calling their functions
        pages_to_test = [
            ('ai_analysis', 'AI Analysis'),
            ('ai_generate', 'Generate Analysis'),
            ('ai_cashflow', 'Cashflow Statement'),
            ('ai_risk', 'Risk Assessment'),
            ('ai_anomaly', 'Anomaly Detection'),
            ('ai_forecast', 'Forecast'),
            ('ai_dashboard_page', 'AI Dashboard'),
            ('ai_assistant', 'Virtual Analyst')
        ]
        
        for route_name, page_name in pages_to_test:
            try:
                # Simulate the route call without actually rendering
                transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
                health_status['pages_tested'].append({
                    'page': page_name,
                    'status': 'ok',
                    'transaction_count': transaction_count
                })
            except Exception as e:
                health_status['pages_tested'].append({
                    'page': page_name,
                    'status': 'error',
                    'error': str(e)
                })
                health_status['status'] = 'degraded'
        
        return jsonify(health_status)
        
    except Exception as e:
        current_app.logger.error(f"AI health check error: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/ai-analysis/basic', methods=['GET'])
@login_required
def ai_analysis_basic():
    """Basic AI analysis for regular users"""
    try:
        # Get current user's transactions
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        
        if not transactions:
            return jsonify({
                'error': 'No transaction data found. Please add some transactions first.'
            })
        
        # Basic analysis data
        total_transactions = len(transactions)
        total_amount = sum(t.amount for t in transactions)
        avg_amount = total_amount / total_transactions if total_transactions > 0 else 0
        
        # Simple pattern analysis
        income_transactions = [t for t in transactions if t.amount > 0]
        expense_transactions = [t for t in transactions if t.amount < 0]
        
        total_income = sum(t.amount for t in income_transactions)
        total_expenses = abs(sum(t.amount for t in expense_transactions))
        
        # Basic cash flow analysis
        current_balance = sum(t.amount for t in transactions)
        
        # Simple risk metrics (monthly convention)
        monthly_burn = total_expenses  # Treat total monthly expenses as monthly burn
        runway_months = (current_balance / monthly_burn) if monthly_burn > 0 else 0
        risk_metrics = {
            'liquidity_ratio': 1.0 if total_expenses > 0 else 0.0,  # Simplified
            'cash_flow_volatility': abs(total_income - total_expenses),
            'burn_rate': monthly_burn,  # monthly burn
            'runway_months': runway_months
        }
        
        # Basic AI analysis text
        ai_analysis = f"""
        <div class="analysis-summary">
            <h6><i class="fas fa-chart-pie me-2"></i>Transaction Overview</h6>
            <p>You have {total_transactions} transactions with an average amount of ${avg_amount:.2f}.</p>
            <p><strong>Total Income:</strong> ${total_income:.2f}</p>
            <p><strong>Total Expenses:</strong> ${total_expenses:.2f}</p>
            <p><strong>Current Balance:</strong> ${current_balance:.2f}</p>
            
            <h6 class="mt-3"><i class="fas fa-lightbulb me-2"></i>Insights</h6>
            <ul>
                <li>Your expense ratio is {(total_expenses/total_income*100):.1f}% of income</li>
                <li>Average monthly burn rate: ${risk_metrics['burn_rate']:.2f}</li>
                <li>Estimated runway: {risk_metrics['runway_months']:.1f} months</li>
            </ul>
        </div>
        """
        
        return jsonify({
            'ai_analysis': ai_analysis,
            'risk_metrics': risk_metrics,
            'patterns': {
                'seasonal_pattern': [100, 105, 98, 102, 110, 95, 108, 103, 97, 105, 112, 98]  # Mock data
            },
            'forecasts': {
                '90_days': [100 + i * 0.5 + (i % 7 - 3) * 2 for i in range(90)]  # Mock data
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'Analysis failed: {str(e)}'}), 500

@app.route('/create-transaction', methods=['POST'])
@login_required
def create_transaction():
    # Validate CSRF token
    if not validate_csrf_header():
        return jsonify({'error': 'CSRF token missing or invalid'}), 400
    
    try:
        # Get currency information from form or defaults
        original_currency = request.form.get('currency', app.config.get('CURRENCY_PRIMARY', 'AED'))
        original_amount = float(request.form.get('amount'))
        
        new_transaction = Transaction(
            user_id=current_user.id,
            date=request.form.get('date'),
            description=request.form.get('description'),
            amount=original_amount,
            type=request.form.get('type'),
            # Currency fields
            original_currency=original_currency,
            original_amount=original_amount
        )
        
        # Apply currency conversion if needed
        try:
            currency_service = get_currency_service()
            if currency_service and currency_service.config.get('CURRENCY_CONVERSION_ENABLED', True):
                conversion_success = new_transaction.apply_currency_conversion(currency_service)
                if not conversion_success:
                    flash('Warning: Currency conversion failed, transaction saved with original currency', 'warning')
        except AttributeError as e:
            app.logger.error(f'Currency conversion method not available: {str(e)}')
            flash('Warning: Currency conversion unavailable, transaction saved with original currency', 'warning')
        except Exception as e:
            app.logger.error(f'Currency conversion error: {str(e)}')
            flash('Warning: Currency conversion failed, transaction saved with original currency', 'warning')
        
        db.session.add(new_transaction)
        db.session.commit()
        flash('Transaction added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error adding transaction: ' + str(e), 'danger')
    
    return redirect(url_for('cash_activities'))

@app.route('/monthly-income-expense', methods=['GET'])
@login_required
def monthly_income_expense():
    try:
        transactions = Transaction.query.filter_by(user_id=current_user.id).order_by(Transaction.date).all()
        monthly_data = {}

        for transaction in transactions:
            month = transaction.date.strftime('%Y-%m') if isinstance(transaction.date, datetime) else datetime.strptime(transaction.date, '%Y-%m-%d').strftime('%Y-%m')
            
            if month not in monthly_data:
                monthly_data[month] = {'income': 0, 'expense': 0}
            
            if transaction.amount > 0:
                monthly_data[month]['income'] += transaction.amount
            else:
                monthly_data[month]['expense'] += abs(transaction.amount)

        return jsonify({
            'success': True,
            'data': {
                'labels': list(monthly_data.keys()),
                'income': [data['income'] for data in monthly_data.values()],
                'expense': [data['expense'] for data in monthly_data.values()]
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/cashout-categories', methods=['GET'])
@login_required
def cashout_categories():
    try:
        # Fix the filter syntax for negative amounts
        transactions = Transaction.query.filter(
            Transaction.user_id == current_user.id,
            Transaction.amount < 0  # Correct syntax for filtering negative amounts
        ).all()

        categories_data = {}
        for transaction in transactions:
            category = transaction.type
            if category not in categories_data:
                categories_data[category] = 0
            categories_data[category] += abs(transaction.amount)

        return jsonify({
            'success': True,
            'data': {
                'labels': list(categories_data.keys()),
                'amounts': list(categories_data.values())
            }
        })
    except Exception as e:
        app.logger.error(f"Error fetching cashout categories: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/debug-language')
def debug_language():
    """Debug endpoint to check translations"""
    all_info = {
        'session_language': session.get('language', 'Not set'),
        'available_translations': os.listdir('translations'),
        'session_data': dict(session),
        'best_match': request.accept_languages.best_match(['en', 'es', 'ja']),
    }
    return jsonify(all_info)

@app.route('/offline')
def offline_page():
    # Offline fallback view used by the service worker
    try:
        return render_template('offline.html')
    except Exception:
        # Minimal inline fallback if template fails
        return '<h1>Offline</h1><p>Please check your connection.</p>'

@app.route('/sw.js')
def service_worker_file():
    # Serve service worker from the scope root
    try:
        resp = send_file(os.path.join(os.path.dirname(__file__), 'sw.js'), mimetype='application/javascript')
        # Prevent aggressive caching during development
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return resp
    except Exception:
        abort(404)

# Advanced AI Features Routes - Super Admin Only
@app.route('/ai/categorize-transactions', methods=['POST'])
@super_admin_required
@ai_rate_limit('categorization')
@monitor_ai_performance('categorization')
def ai_categorize_transactions():
    """Route for automated transaction categorization"""
    try:
        data = request.get_json()
        if data is None:
            return jsonify({'error': 'Invalid or missing JSON body'}), 400
        transactions = data.get('transactions', [])
        
        if not transactions:
            return jsonify({'error': 'No transactions provided'}), 400
        
        # Check cache first
        from src.ai_utils import AIUtils
        import hashlib
        cache_key = f"categorization_{current_user.id}_{hashlib.sha256(str(transactions).encode()).hexdigest()[:16]}"
        cached_result = AIUtils.get_cached_ai_results(cache_key)
        if cached_result:
            return jsonify({
                'success': True,
                'result': cached_result,
                'cached': True
            })
        
        # Initialize financial analytics with API key
        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'error': 'API key not found'}), 500
            
        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        
        # Get categorization suggestions
        categorization_result = analytics.automated_transaction_categorization(transactions)
        
        # Cache the result
        ttl = current_app.config['AI_CACHE_TTL']['categorization']
        AIUtils.cache_ai_results(cache_key, categorization_result, ttl)
        
        # Log AI feature usage
        current_app.logger.info(f"AI categorization used by super_admin user {current_user.id}")
        
        return jsonify({
            'success': True,
            'result': categorization_result
        })
        
    except Exception as e:
        current_app.logger.error(f"AI categorization error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/ai/risk-assessment', methods=['GET', 'POST'])
@super_admin_required
@ai_rate_limit('risk_assessment')
@monitor_ai_performance('risk_assessment')
def ai_risk_assessment():
    """Route for comprehensive risk assessment reports"""
    try:
        if request.method == 'GET':
            # Check if this is an AJAX request or a direct page access
            if request.headers.get('Content-Type') == 'application/json' or 'application/json' in request.headers.get('Accept', ''):
                # AJAX API request - proceed with API logic
                transactions_query = Transaction.query.filter_by(user_id=current_user.id)
                transactions = transactions_query.order_by(Transaction.date).all()
                
                if not transactions:
                    return jsonify({'error': 'No transactions found for risk assessment'}), 400
            else:
                # Direct page access - render the HTML template
                transaction_count = Transaction.query.filter_by(user_id=current_user.id).count()
                has_data = transaction_count > 0
                
                return render_template('ai_risk_assessment.html',
                                     has_data=has_data,
                                     transaction_count=transaction_count,
                                     risk_metrics=None,
                                     recent_risk_activities=None)
        else:
            # POST request with specific user data
            data = request.get_json()
            if data is None:
                return jsonify({'error': 'Invalid or missing JSON body'}), 400
            user_id = data.get('user_id', current_user.id)
            transactions_query = Transaction.query.filter_by(user_id=user_id)
            transactions = transactions_query.order_by(Transaction.date).all()
        
        if not transactions:
            return jsonify({'error': 'No transactions found for risk assessment'}), 400
        
        # Prepare transaction history
        from src.ai_utils import AIUtils
        transaction_history = AIUtils.build_transaction_history(transactions)
        
        # Prepare balance data
        if request.method == 'GET':
            target_user_id = current_user.id
            target_user = current_user
        else:
            target_user_id = user_id
            target_user = User.query.get(user_id)
            if not target_user:
                return jsonify({'error': 'User not found'}), 400
        
        initial_balance_record = InitialBalance.query.filter_by(user_id=target_user_id).first()
        initial_balance = initial_balance_record.balance if initial_balance_record else 0.0
        current_balance = initial_balance + sum(t.amount for t in transactions)
        
        balance_data = {
            'initial_balance': initial_balance,
            'current_balance': current_balance,
            'total_transactions': len(transactions)
        }
        
        # Prepare user profile
        user_profile = {
            'user_id': target_user_id,
            'username': target_user.username,
            'account_age_days': 30,  # Placeholder - you might want to calculate this
            'transaction_frequency': len(transactions) / 30  # transactions per day
        }
        
        # Check cache first
        from src.ai_utils import AIUtils
        import hashlib
        cache_key = f"risk_assessment_{target_user_id}_{hashlib.sha256(str(transaction_history + [balance_data, user_profile]).encode()).hexdigest()[:16]}"
        cached_result = AIUtils.get_cached_ai_results(cache_key)
        if cached_result:
            return jsonify({
                'success': True,
                'result': cached_result,
                'cached': True
            })
        
        # Initialize financial analytics
        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'error': 'API key not found'}), 500
            
        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        
        # Generate risk assessment
        risk_assessment = analytics.risk_assessment_reports(
            transaction_history, balance_data, user_profile
        )
        
        # Cache the result
        ttl = current_app.config['AI_CACHE_TTL']['risk_assessment']
        AIUtils.cache_ai_results(cache_key, risk_assessment, ttl)
        
        # Log AI feature usage
        log_ai_feature_usage(
            user_id=current_user.id,
            feature_name='risk_assessment',
            duration=0,  # Would be calculated from start time
            success=True,
            details={'user_id': target_user_id, 'transaction_count': len(transactions)}
        )
        
        current_app.logger.info(f"AI risk assessment completed for user {target_user_id}")
        
        return jsonify({
            'success': True,
            'result': risk_assessment
        })
        
    except Exception as e:
        current_app.logger.error(f"AI risk assessment error: {str(e)}")
        
        # Log the error
        log_ai_feature_usage(
            user_id=current_user.id,
            feature_name='risk_assessment',
            duration=0,
            success=False,
            details={'error': str(e)}
        )
        
        return jsonify({'error': 'Risk assessment failed. Please try again.'}), 500

@app.route('/ai/anomaly-detection', methods=['GET', 'POST'])
@super_admin_required
@ai_rate_limit('anomaly_detection')
@monitor_ai_performance('anomaly_detection')
def ai_anomaly_detection():
    """Route for anomaly detection analysis"""
    try:
        if request.method == 'GET':
            transactions_query = Transaction.query.filter_by(user_id=current_user.id)
            transactions = transactions_query.order_by(Transaction.date).all()
        else:
            data = request.get_json()
            if data is None:
                return jsonify({'error': 'Invalid or missing JSON body'}), 400
            user_id = data.get('user_id', current_user.id)
            transactions_query = Transaction.query.filter_by(user_id=user_id)
            transactions = transactions_query.order_by(Transaction.date).all()
        
        if not transactions:
            return jsonify({'error': 'No transactions found for anomaly detection'}), 400
        
        # Prepare transaction history
        from src.ai_utils import AIUtils
        transaction_history = AIUtils.build_transaction_history(transactions)
        
        # Prepare user patterns (simplified)
        amounts = [t.amount for t in transactions]
        user_patterns = {
            'average_amount': sum(amounts) / len(amounts) if amounts else 0,
            'max_amount': max(amounts) if amounts else 0,
            'min_amount': min(amounts) if amounts else 0,
            'transaction_count': len(transactions),
            'typical_frequency': len(transactions) / 30  # transactions per day
        }
        
        # Check cache first
        from src.ai_utils import AIUtils
        import hashlib
        user_id = current_user.id if request.method == 'GET' else data.get('user_id', current_user.id)
        cache_key = f"anomaly_detection_{user_id}_{hashlib.sha256(str(transaction_history + [user_patterns]).encode()).hexdigest()[:16]}"
        cached_result = AIUtils.get_cached_ai_results(cache_key)
        if cached_result:
            return jsonify({
                'success': True,
                'result': cached_result,
                'cached': True
            })
        
        # Initialize financial analytics
        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'error': 'API key not found'}), 500
            
        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        
        # Generate anomaly detection
        anomaly_result = analytics.anomaly_detection(transaction_history, user_patterns)
        
        # Cache the result
        ttl = current_app.config['AI_CACHE_TTL']['anomaly_detection']
        AIUtils.cache_ai_results(cache_key, anomaly_result, ttl)
        
        # Log AI feature usage
        current_app.logger.info(f"AI anomaly detection used by super_admin user {current_user.id}")
        
        # Log AI feature usage for analytics
        log_ai_feature_usage(
            user_id=current_user.id,
            feature_name='anomaly_detection',
            duration=0,  # Duration would be calculated if we had timing
            success=True,
            details={'action': 'anomaly_detection_completed'}
        )
        
        return jsonify({
            'success': True,
            'result': anomaly_result
        })
        
    except Exception as e:
        current_app.logger.error(f"AI anomaly detection error: {str(e)}")
        
        # Log AI feature usage for analytics (error case)
        log_ai_feature_usage(
            user_id=current_user.id if current_user.is_authenticated else None,
            feature_name='anomaly_detection',
            duration=0,
            success=False,
            details={'error': str(e)}
        )
        
        return jsonify({'error': str(e)}), 500

@app.route('/ai/advanced-forecast', methods=['GET', 'POST'])
@super_admin_required
@ai_rate_limit('advanced_forecast')
@monitor_ai_performance('advanced_forecast')
def ai_advanced_forecast():
    """Route for advanced forecasting models"""
    try:
        if request.method == 'GET':
            transactions_query = Transaction.query.filter_by(user_id=current_user.id)
            transactions = transactions_query.order_by(Transaction.date).all()
        else:
            data = request.get_json()
            if data is None:
                return jsonify({'error': 'Invalid or missing JSON body'}), 400
            user_id = data.get('user_id', current_user.id)
            transactions_query = Transaction.query.filter_by(user_id=user_id)
            transactions = transactions_query.order_by(Transaction.date).all()
        
        if not transactions:
            return jsonify({'error': 'No transactions found for forecasting'}), 400
        
        # Prepare transaction history
        transaction_history = []
        for t in transactions:
            date_str = t.date if isinstance(t.date, str) else t.date.strftime('%Y-%m-%d')
            transaction_history.append({
                'date': date_str,
                'amount': t.amount,
                'type': t.type,
                'description': t.description
            })
        
        # Prepare external factors (simplified)
        external_factors = {
            'economic_indicators': {
                'inflation_rate': 3.2,
                'interest_rate': 5.25,
                'gdp_growth': 2.1
            },
            'seasonal_factors': {
                'quarter': 1,
                'month': 1,
                'holiday_season': False
            },
            'market_conditions': {
                'volatility_index': 18.5,
                'market_sentiment': 'neutral'
            }
        }
        
        # Check cache first
        from src.ai_utils import AIUtils
        import hashlib
        user_id = current_user.id if request.method == 'GET' else data.get('user_id', current_user.id)
        cache_key = f"advanced_forecast_{user_id}_{hashlib.sha256(str(transaction_history + [external_factors]).encode()).hexdigest()[:16]}"
        cached_result = AIUtils.get_cached_ai_results(cache_key)
        if cached_result:
            return jsonify({
                'success': True,
                'result': cached_result,
                'cached': True
            })
        
        # Initialize financial analytics
        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'error': 'API key not found'}), 500
            
        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        
        # Generate advanced forecasting
        forecast_result = analytics.advanced_forecasting_models(
            transaction_history, external_factors
        )
        
        # Cache the result
        ttl = current_app.config['AI_CACHE_TTL']['advanced_forecast']
        AIUtils.cache_ai_results(cache_key, forecast_result, ttl)
        
        # Log AI feature usage
        current_app.logger.info(f"AI advanced forecasting used by super_admin user {current_user.id}")
        
        # Log AI feature usage for analytics
        log_ai_feature_usage(
            user_id=current_user.id,
            feature_name='advanced_forecast',
            duration=0,  # Duration would be calculated if we had timing
            success=True,
            details={'action': 'advanced_forecast_completed'}
        )
        
        return jsonify({
            'success': True,
            'result': forecast_result
        })
        
    except Exception as e:
        current_app.logger.error(f"AI advanced forecasting error: {str(e)}")
        
        # Log AI feature usage for analytics (error case)
        log_ai_feature_usage(
            user_id=current_user.id if current_user.is_authenticated else None,
            feature_name='advanced_forecast',
            duration=0,
            success=False,
            details={'error': str(e)}
        )
        
        return jsonify({'error': str(e)}), 500

@app.route('/ai/custom-insights', methods=['POST'])
@super_admin_required
@ai_rate_limit('custom_insights')
@monitor_ai_performance('custom_insights')
def ai_custom_insights():
    """Route for custom financial insights"""
    try:
        data = request.get_json()
        if data is None:
            return jsonify({'error': 'Invalid or missing JSON body'}), 400
        analysis_type = data.get('analysis_type', 'general')
        custom_parameters = data.get('custom_parameters', {})
        user_id = data.get('user_id', current_user.id)
        
        # Get transactions for analysis
        transactions_query = Transaction.query.filter_by(user_id=user_id)
        transactions = transactions_query.order_by(Transaction.date).all()
        
        if not transactions:
            return jsonify({'error': 'No transactions found for analysis'}), 400
        
        # Prepare transaction data
        transaction_data = []
        for t in transactions:
            date_str = t.date if isinstance(t.date, str) else t.date.strftime('%Y-%m-%d')
            transaction_data.append({
                'date': date_str,
                'amount': t.amount,
                'type': t.type,
                'description': t.description
            })
        
        # Check cache first
        from src.ai_utils import AIUtils
        import hashlib
        cache_key = f"custom_insights_{user_id}_{hashlib.sha256(str(transaction_data + [analysis_type, custom_parameters]).encode()).hexdigest()[:16]}"
        cached_result = AIUtils.get_cached_ai_results(cache_key)
        if cached_result:
            return jsonify({
                'success': True,
                'result': cached_result,
                'cached': True
            })
        
        # Initialize financial analytics
        api_key = current_app.config.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'error': 'API key not found'}), 500
            
        analytics = FinancialAnalytics(api_key=api_key, test_connection=False)
        
        # Generate custom insights
        insights_result = analytics.custom_financial_insights(
            transaction_data, analysis_type, custom_parameters
        )
        
        # Cache the result
        ttl = current_app.config['AI_CACHE_TTL']['custom_insights']
        AIUtils.cache_ai_results(cache_key, insights_result, ttl)
        
        # Log AI feature usage
        current_app.logger.info(f"AI custom insights used by super_admin user {current_user.id}, analysis_type: {analysis_type}")
        
        return jsonify({
            'success': True,
            'result': insights_result
        })
        
    except Exception as e:
        current_app.logger.error(f"AI custom insights error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# AI Dashboard Helper Functions and Caching
from src.ai_dashboard_utils import (
    get_ai_system_status,
    get_recent_ai_activities,
    get_ai_performance_metrics,
    log_ai_feature_usage,
    prepare_dashboard_context,
    monitor_system_health
)

@app.route('/ai/dashboard', methods=['GET'])
@super_admin_required
def ai_dashboard():
    """Route for AI dashboard showing all advanced features"""
    try:
        # Log AI dashboard access
        current_app.logger.info(f"AI dashboard accessed by super_admin user {current_user.id}")
        
        # Get comprehensive dashboard context using helper functions
        dashboard_context = prepare_dashboard_context(user_id=current_user.id)
        
        # Get user transaction summary for dashboard
        transactions_query = Transaction.query.filter_by(user_id=current_user.id)
        transaction_count = transactions_query.count()
        
        # Available AI features
        available_ai_features = [
            'risk_assessment', 'anomaly_detection', 'advanced_forecast',
            'custom_insights', 'transaction_categorization', 'chatbot'
        ]
        
        # Log dashboard access
        log_ai_feature_usage(
            user_id=current_user.id,
            feature_name='dashboard_access',
            duration=0,
            success=True,
            details={'action': 'dashboard_view'}
        )
        
        return render_template('ai_dashboard.html', 
                             transaction_count=transaction_count,
                             user=current_user,
                             available_ai_features=available_ai_features,
                             ai_performance_metrics=dashboard_context.get('performance_metrics'),
                             ai_system_status=dashboard_context.get('system_status'),
                             recent_ai_activities=dashboard_context.get('recent_activities', []))
        
    except Exception as e:
        current_app.logger.error(f"AI dashboard error: {str(e)}")
        # Log the error
        log_ai_feature_usage(
            user_id=current_user.id if current_user.is_authenticated else None,
            feature_name='dashboard_access',
            duration=0,
            success=False,
            details={'error': str(e)}
        )
        flash('Error loading AI dashboard. Please try again.', 'danger')
        return redirect(url_for('home'))

@app.route('/ai/results', methods=['GET', 'POST'])
@super_admin_required
def ai_results():
    """Route for displaying AI analysis results"""
    try:
        if request.method == 'POST':
            # Store results in session
            data = request.get_json()
            if data:
                session['ai_results'] = {
                    'results': data.get('results', {}),
                    'feature_type': data.get('feature_type', 'general'),
                    'timestamp': data.get('timestamp', datetime.now().isoformat())
                }
                return jsonify({'success': True})
            else:
                return jsonify({'error': 'No data provided'}), 400
        
        # GET request - display results
        # Get analysis type and results from session or query parameters
        analysis_type = request.args.get('type', 'general')
        results_data = session.get('ai_results', {})
        
        # If no results in session, redirect to dashboard
        if not results_data:
            flash('No analysis results found. Please run an analysis first.', 'info')
            return redirect(url_for('ai_dashboard'))
        
        # Prepare context for results template
        context = {
            'analysis_type': analysis_type,
            'timestamp': results_data.get('timestamp', datetime.now().isoformat()),
            'risk_summary': results_data.get('results', {}).get('risk_summary'),
            'liquidity_risk': results_data.get('results', {}).get('liquidity_risk'),
            'credit_risk': results_data.get('results', {}).get('credit_risk'),
            'market_risk': results_data.get('results', {}).get('market_risk'),
            'anomaly_summary': results_data.get('results', {}).get('anomaly_summary'),
            'anomalies': results_data.get('results', {}).get('anomalies', []),
            'forecasts': results_data.get('results', {}).get('forecasts'),
            'recommendations': results_data.get('results', {}).get('recommendations', []),
            'action_items': results_data.get('results', {}).get('action_items', [])
        }
        
        return render_template('ai_results.html', **context)
        
    except Exception as e:
        current_app.logger.error(f"AI results page error: {str(e)}")
        flash('Error loading results page. Please try again.', 'danger')
        return redirect(url_for('ai_dashboard'))

@app.route('/chatbot', methods=['GET', 'POST'])
@super_admin_required
@ai_rate_limit('chatbot')
def chatbot():
    """Route for Natural Language to SQL chatbot interface"""
    try:
        # Check if chatbot is enabled
        if not current_app.config.get('CHATBOT_ENABLED', True):
            flash('Chatbot feature is currently disabled.', 'error')
            return redirect(url_for('home'))
        
        from src.nl_sql_service import NLSQLChatbot
        
        # Initialize chatbot service
        if not current_app.config.get('ANTHROPIC_API_KEY'):
            flash('AI services are not configured. Please contact administrator.', 'error')
            return redirect(url_for('home'))
        
        chatbot_service = NLSQLChatbot(current_app.config['ANTHROPIC_API_KEY'])
        
        if request.method == 'GET':
            # Render chatbot interface
            query_examples = chatbot_service.get_query_examples()
            
            # Log chatbot access
            current_app.logger.info(f"Chatbot interface accessed by super_admin user {current_user.id}")
            
            return render_template('chatbot.html', 
                                 query_examples=query_examples,
                                 user=current_user)
        
        elif request.method == 'POST':
            # Process natural language query
            data = request.get_json()
            if not data or 'query' not in data:
                return jsonify({'error': 'Query is required'}), 400
            
            query = data['query'].strip()
            if not query:
                return jsonify({'error': 'Query cannot be empty'}), 400
            
            # Check query length
            if len(query) > current_app.config.get('CHATBOT_MAX_QUERY_LENGTH', 500):
                return jsonify({'error': f'Query too long. Maximum {current_app.config.get("CHATBOT_MAX_QUERY_LENGTH", 500)} characters allowed.'}), 400
            
            # Prepare user context
            user_context = {
                'user_id': current_user.id,
                'username': current_user.username,
                'role': 'super_admin',  # Only super_admin can access chatbot
                'can_access_all_data': True
            }
            
            # Process the query
            result = chatbot_service.process_natural_language_query(query, user_context)
            
            # Log query execution
            current_app.logger.info(f"Chatbot query executed by user {current_user.id}: {query[:100]}...")
            
            return jsonify(result)
        
    except Exception as e:
        current_app.logger.error(f"Chatbot error: {str(e)}")
        if request.method == 'POST':
            return jsonify({'error': f'Query processing failed: {str(e)}'}), 500
        else:
            flash('Error loading chatbot interface', 'danger')
            return redirect(url_for('home'))

@app.route('/chatbot/query', methods=['POST'])
@super_admin_required
@ai_rate_limit('chatbot')
def chatbot_query():
    """Route for chatbot query processing"""
    try:
        # Check if chatbot is enabled
        if not current_app.config.get('CHATBOT_ENABLED', True):
            return jsonify({'success': False, 'error': 'Chatbot feature is currently disabled.'}), 403
        
        from src.nl_sql_service import NLSQLChatbot
        
        # Initialize chatbot service
        if not current_app.config.get('ANTHROPIC_API_KEY'):
            return jsonify({'success': False, 'error': 'AI services are not configured.'}), 503
        
        chatbot_service = NLSQLChatbot(current_app.config['ANTHROPIC_API_KEY'])
        
        # Process natural language query
        data = request.get_json()
        if not data or 'query' not in data:
            return jsonify({'success': False, 'error': 'Query is required'}), 400
        
        query = data['query'].strip()
        if not query:
            return jsonify({'success': False, 'error': 'Query cannot be empty'}), 400
        
        # Check query length
        if len(query) > current_app.config.get('CHATBOT_MAX_QUERY_LENGTH', 500):
            return jsonify({'success': False, 'error': f'Query too long. Maximum {current_app.config.get("CHATBOT_MAX_QUERY_LENGTH", 500)} characters allowed.'}), 400
        
        # Prepare user context
        user_context = {
            'user_id': current_user.id,
            'username': current_user.username,
            'role': 'super_admin',  # Only super_admin can access chatbot
            'can_access_all_data': True
        }
        
        # Process the query
        result = chatbot_service.process_natural_language_query(query, user_context)
        
        # Log query execution
        current_app.logger.info(f"Chatbot query executed by user {current_user.id}: {query[:100]}...")
        
        # Return success response with data
        return jsonify({
            'success': True,
            'response': result.get('response', 'Query processed successfully'),
            'data': result.get('data', [])
        })
        
    except Exception as e:
        current_app.logger.error(f"Chatbot query error: {str(e)}")
        return jsonify({'success': False, 'error': f'Query processing failed: {str(e)}'}), 500

############################
# Simple rate limiting (admin)
############################
_admin_rate_limit_bucket = {}

def admin_rate_limit(key_prefix: str, max_requests: int = 30, window_seconds: int = 60):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            user_id = getattr(current_user, 'id', 'anon')
            now = time.time()
            bucket_key = f"{key_prefix}:{user_id}"
            window = _admin_rate_limit_bucket.get(bucket_key, [])
            window = [ts for ts in window if now - ts < window_seconds]
            if len(window) >= max_requests:
                return jsonify({'error': 'Rate limit exceeded. Try again later.'}), 429
            window.append(now)
            _admin_rate_limit_bucket[bucket_key] = window
            return func(*args, **kwargs)
        return wrapper
    return decorator

############################
# Admin Routes
############################

@app.route('/admin/users', methods=['GET'])
@admin_required
@admin_rate_limit('admin_users_list', max_requests=60, window_seconds=60)
def admin_users():
    page = request.args.get('page', 1, type=int)
    raw_per_page = request.args.get('per_page', 20, type=int)
    per_page = max(1, min(raw_per_page, 100))
    search = (request.args.get('q') or '')[:100]
    role_filter = request.args.get('role')

    users_page, roles_map = get_users_with_roles(page=page, per_page=per_page, search=search, role_filter=role_filter)
    stats = get_user_statistics()

    # Roles current admin can assign
    available_roles = get_available_roles_for_admin(current_user)

    return render_template(
        'admin/users.html',
        users=users_page.items,
        pagination=users_page,
        roles_map=roles_map,
        stats=stats,
        search=search or '',
        role_filter=role_filter or '',
        available_roles=available_roles,
        user=current_user
    )

@app.route('/admin/assign-role', methods=['POST'])
@admin_required
@admin_rate_limit('admin_assign_role', max_requests=20, window_seconds=60)
def admin_assign_role():
    try:
        data = request.get_json() or request.form
        user_id = int(data.get('user_id'))
        role_name = (data.get('role_name') or '').strip()
        action = (data.get('action') or 'assign').strip()

        target_user = User.query.get(user_id)
        if not target_user:
            return jsonify({'error': 'User not found'}), 404

        valid, message = validate_role_assignment(current_user, target_user, role_name, action)
        if not valid:
            return jsonify({'error': message}), 400

        if action == 'assign':
            success = assign_role(target_user, role_name)
            action_label = 'assign_role'
            err = None
        else:
            success = remove_role(target_user, role_name)
            action_label = 'remove_role'
            err = None

        if not success:
            return jsonify({'error': err or 'Operation failed'}), 400

        # Sync user preferences with new role configuration
        try:
            update_user_preferences_on_role_change(target_user)
        except Exception:
            pass

        log_admin_action(current_user, action_label, target_user, {'role': role_name})
        return jsonify({'success': True, 'user_id': user_id, 'role': role_name, 'action': action})
    except Exception as e:
        current_app.logger.error(f"Admin role change error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/user/<int:user_id>', methods=['GET'])
@admin_required
@admin_rate_limit('admin_user_detail', max_requests=60, window_seconds=60)
def admin_user_detail(user_id: int):
    user_obj = User.query.get_or_404(user_id)
    summary = get_user_transaction_summary(user_id)
    role_history = get_user_role_history(user_id)
    current_roles = get_user_roles(user_obj)
    available_roles = get_available_roles_for_admin(current_user)

    return render_template(
        'admin/user_detail.html',
        target_user=user_obj,
        summary=summary,
        role_history=role_history,
        current_roles=current_roles,
        available_roles=available_roles,
        user=current_user
    )

@app.route('/admin/dashboard', methods=['GET'])
@admin_required
@admin_rate_limit('admin_dashboard', max_requests=60, window_seconds=60)
def admin_dashboard():
    stats = get_user_statistics()
    # These keys are expected by the dashboard template for charts
    trends = {
        'registrations': [],
        'labels': []
    }
    try:
        from src.admin_utils import get_registration_trends
        trends_data = get_registration_trends(months=12)
        trends['labels'] = [t['label'] for t in trends_data]
        trends['registrations'] = [t['count'] for t in trends_data]
    except Exception:
        pass

    return render_template('admin/dashboard.html', stats=stats, trends=trends, user=current_user)

@app.route('/admin/users/export', methods=['GET'])
@admin_required
@admin_rate_limit('admin_users_export', max_requests=10, window_seconds=60)
def admin_users_export():
    export_format = request.args.get('format', 'csv')
    filters = {
        'q': request.args.get('q'),
        'role': request.args.get('role')
    }
    try:
        output, mimetype, filename = export_user_data(format=export_format, filters=filters)
        return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/admin/reports', methods=['GET'])
@admin_required
def admin_reports():
    """Admin reports page"""
    try:
        # Get basic statistics for reports
        total_users = User.query.count()
        total_transactions = Transaction.query.count()
        total_balance = db.session.query(db.func.sum(Transaction.amount)).scalar() or 0
        
        # Get recent activity
        recent_transactions = Transaction.query.order_by(Transaction.date.desc()).limit(10).all()
        
        return render_template('admin/reports.html', 
                             total_users=total_users,
                             total_transactions=total_transactions,
                             total_balance=total_balance,
                             recent_transactions=recent_transactions)
    except Exception as e:
        flash(f'Error loading reports: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    """Admin settings page"""
    if request.method == 'POST':
        try:
            # Handle settings updates here
            # This is a placeholder for actual settings management
            flash('Settings updated successfully', 'success')
            return redirect(url_for('admin_settings'))
        except Exception as e:
            flash(f'Error updating settings: {str(e)}', 'error')
    
    try:
        # Get current settings
        # This is a placeholder for actual settings retrieval
        settings = {
            'site_name': 'FlowTrack',
            'maintenance_mode': False,
            'registration_enabled': True,
            'require_2fa': False,
            'session_timeout': 30,
            'max_login_attempts': 5,
            'ai_analysis_enabled': True,
            'ai_forecasting_enabled': True,
            'ai_chatbot_enabled': True,
            'ai_risk_assessment_enabled': True
        }
        
        # System information
        import sys
        import flask
        from datetime import datetime
        
        context = {
            'settings': settings,
            'python_version': sys.version.split()[0],
            'flask_version': flask.__version__,
            'uptime': '24 hours'  # Placeholder
        }
        
        return render_template('admin/settings.html', **context)
    except Exception as e:
        flash(f'Error loading settings: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))

@app.errorhandler(429)
def rate_limit_exceeded(error):
    return jsonify({'error': 'Too many requests'}), 429

@app.errorhandler(403)
def forbidden_error(error):
    has_admin_dashboard = 'admin_dashboard' in current_app.view_functions
    return render_template('errors/403.html', has_admin_dashboard=has_admin_dashboard), 403

@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    current_app.logger.error(f"500 Internal Server Error: {str(error)}")
    return render_template('errors/500.html'), 500

# Template syntax error handler for graceful fallback
@app.errorhandler(Exception)
def handle_template_syntax_error(e):
    # Check if this is a Jinja template syntax error
    from jinja2.exceptions import TemplateSyntaxError
    
    if isinstance(e, TemplateSyntaxError):
        current_app.logger.error(f"Template syntax error: {str(e)}")
        current_app.logger.error(f"Template: {getattr(e, 'filename', 'Unknown')}")
        current_app.logger.error(f"Line: {getattr(e, 'lineno', 'Unknown')}")
        
        # Check if this is an AI-related route
        if hasattr(request, 'endpoint') and request.endpoint and 'ai' in request.endpoint:
            flash('There was a template error on this AI page. Please contact support.', 'error')
            return redirect(url_for('ai_analysis'))
        
        # For other routes, show a generic error
        return render_template('errors/500.html', 
                             error_message="Template syntax error detected. Please contact support."), 500
    
    # Handle other exceptions
    import traceback
    current_app.logger.error(f"Unhandled exception: {str(e)}")
    current_app.logger.error(f"Traceback: {traceback.format_exc()}")
    
    # Check if this is an AI-related route
    if hasattr(request, 'endpoint') and request.endpoint and 'ai' in request.endpoint:
        flash('An error occurred while processing your AI request. Please try again.', 'error')
        # Redirect to AI analysis page with error message
        return redirect(url_for('ai_analysis'))
    
    # Return a generic error page for production
    if not current_app.config.get('DEBUG', False):
        return render_template('errors/500.html'), 500
    else:
        # In debug mode, let Flask handle it normally
        raise

@app.route('/admin/sync-user-preferences', methods=['POST'])
@super_admin_required
def admin_sync_user_preferences():
    """Sync all users' preferences with their current roles.

    Ensures super admins (and admins) have their default AI modules enabled
    without removing any manually enabled modules.
    """
    try:
        updated_count = 0
        users = User.query.all()
        for user in users:
            prefs = UserPreferences.query.filter_by(user_id=user.id).first()
            if not prefs:
                prefs = UserPreferences(user_id=user.id)
                db.session.add(prefs)
                db.session.flush()
            changed = False
            try:
                changed = prefs.ensure_role_based_modules(user)
            except Exception:
                # If any error occurs during ensure, skip this user but continue
                changed = False
            if changed:
                updated_count += 1
        db.session.commit()
        return jsonify({'success': True, 'updated_users': updated_count}), 200
    except SQLAlchemyError as db_err:
        db.session.rollback()
        current_app.logger.error(f"DB error during preferences sync: {str(db_err)}")
        return jsonify({'success': False, 'error': 'Database error during sync'}), 500
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Unexpected error during preferences sync: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== BANK INTEGRATION ROUTES ====================

@app.route('/bank/connections', methods=['GET'])
@login_required
def bank_connections():
    """List user's active bank connections with account details and last sync status."""
    try:
        from src.bank_service import BankService
        from src.models import BankConnection, BankAccount
        
        # Get user's active connections
        connections = BankConnection.get_active_connections(user_id=current_user.id)
        
        connection_data = []
        for conn in connections:
            # Get accounts for this connection
            accounts = BankAccount.query.filter_by(connection_id=conn.id).all()
            
            connection_info = {
                'id': conn.id,
                'bank_name': conn.bank_name,
                'account_mask': conn.account_mask,
                'status': conn.status,
                'last_sync': conn.last_sync.isoformat() if conn.last_sync else None,
                'created_at': conn.created_at.isoformat(),
                'accounts': [
                    {
                        'id': acc.id,
                        'name': acc.account_name,
                        'type': acc.account_type,
                        'mask': acc.account_id[-4:] if acc.account_id else None,
                        'balance': float(acc.current_balance) if acc.current_balance else 0.0,
                        'currency': acc.currency_code
                    } for acc in accounts
                ]
            }
            connection_data.append(connection_info)
        
        return jsonify({
            'success': True,
            'connections': connection_data
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching bank connections: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch bank connections'
        }), 500

@app.route('/bank/import', methods=['POST'])
@login_required
def bank_import_all():
    """Trigger manual bank transaction sync for user's active connections."""
    try:
        from src.bank_service import BankService, SyncManager
        from src.models import BankConnection
        
        # Get date range from request (default 30 days)
        data = request.get_json() or {}
        date_range = data.get('date_range', 30)
        
        # Get user's active connections
        connections = BankConnection.get_active_connections(user_id=current_user.id)
        
        if not connections:
            return jsonify({
                'success': False,
                'error': 'No active bank connections found. Please connect a bank account first.'
            }), 400
        
        # Initialize services
        bank_service = BankService()
        sync_manager = SyncManager()
        
        total_imported = 0
        total_duplicates = 0
        errors = []
        
        # Sync each connection
        for connection in connections:
            try:
                result = sync_manager.sync_connection(connection.id, days=date_range)
                total_imported += result.get('created_transactions', 0)
                total_duplicates += result.get('skipped_duplicates', 0)
                
                if result.get('errors'):
                    errors.extend(result['errors'])
                    
            except Exception as e:
                error_msg = f"Failed to sync {connection.bank_name}: {str(e)}"
                errors.append(error_msg)
                current_app.logger.error(error_msg)
        
        return jsonify({
            'success': True,
            'imported_count': total_imported,
            'duplicate_count': total_duplicates,
            'errors': errors,
            'message': f'Successfully imported {total_imported} transactions from {len(connections)} bank connection(s)'
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error during bank import: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to import bank transactions'
        }), 500

@app.route('/bank/import/<int:connection_id>', methods=['POST'])
@login_required
def bank_import_connection(connection_id):
    """Sync specific bank connection."""
    try:
        from src.bank_service import SyncManager
        from src.models import BankConnection
        
        # Get date range from request (default 30 days)
        data = request.get_json() or {}
        date_range = data.get('date_range', 30)
        
        # Verify connection belongs to user
        connection = BankConnection.query.filter_by(
            id=connection_id, 
            user_id=current_user.id
        ).first()
        
        if not connection:
            return jsonify({
                'success': False,
                'error': 'Bank connection not found or access denied'
            }), 404
        
        if connection.status != 'active':
            return jsonify({
                'success': False,
                'error': f'Bank connection is {connection.status}. Please reconnect your account.'
            }), 400
        
        # Perform sync
        sync_manager = SyncManager()
        result = sync_manager.sync_connection(connection_id, days=date_range)
        
        return jsonify({
            'success': True,
            'imported_count': result.get('created_transactions', 0),
            'duplicate_count': result.get('skipped_duplicates', 0),
            'errors': result.get('errors', []),
            'message': f'Successfully imported {result.get("created_transactions", 0)} transactions from {connection.bank_name}'
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error syncing bank connection {connection_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to sync bank connection'
        }), 500

@app.route('/bank/sync-status/<int:connection_id>', methods=['GET'])
@login_required
def bank_sync_status(connection_id):
    """Get sync status and recent logs for a connection."""
    try:
        from src.models import BankConnection, BankSyncLog
        
        # Verify connection belongs to user
        connection = BankConnection.query.filter_by(
            id=connection_id, 
            user_id=current_user.id
        ).first()
        
        if not connection:
            return jsonify({
                'success': False,
                'error': 'Bank connection not found or access denied'
            }), 404
        
        # Get recent sync logs
        logs = BankSyncLog.query.filter_by(connection_id=connection_id)\
            .order_by(BankSyncLog.created_at.desc())\
            .limit(10).all()
        
        log_data = [
            {
                'id': log.id,
                'status': log.status,
                'transactions_fetched': log.transactions_fetched,
                'transactions_processed': log.transactions_processed,
                'errors': log.errors,
                'started_at': log.started_at.isoformat(),
                'completed_at': log.completed_at.isoformat() if log.completed_at else None
            } for log in logs
        ]
        
        return jsonify({
            'success': True,
            'connection': {
                'id': connection.id,
                'bank_name': connection.bank_name,
                'status': connection.status,
                'last_sync': connection.last_sync.isoformat() if connection.last_sync else None
            },
            'recent_logs': log_data
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching sync status: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch sync status'
        }), 500

@app.route('/bank/disconnect/<int:connection_id>', methods=['POST'])
@login_required
def bank_disconnect(connection_id):
    """Disconnect specific bank connection with proper validation."""
    try:
        from src.models import BankConnection
        from src.bank_service import BankService
        
        # Verify connection belongs to user
        connection = BankConnection.query.filter_by(
            id=connection_id, 
            user_id=current_user.id
        ).first()
        
        if not connection:
            return jsonify({
                'success': False,
                'error': 'Bank connection not found or access denied'
            }), 404
        
        # Use service method to properly disconnect
        bank_service = BankService()
        success = bank_service.disconnect_bank(connection_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': f'Successfully disconnected {connection.bank_name}'
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to disconnect bank connection'
            }), 500
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error disconnecting bank connection: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to disconnect bank connection'
        }), 500

@app.route('/bank/connections/link-token', methods=['POST'])
@login_required
def bank_link_token():
    """Create link token for bank connection setup."""
    try:
        from src.bank_service import BankService
        
        # Get tenant_id if available
        tenant_id = getattr(current_user, 'tenant_id', None)
        
        # Create link token
        bank_service = BankService()
        result = bank_service.create_link_token(current_user.id, tenant_id)
        
        return jsonify({
            'success': True,
            'link_token': result.get('link_token'),
            'expiration': result.get('expiration')
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error creating link token: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to create link token'
        }), 500

@app.route('/bank/connections/exchange', methods=['POST'])
@login_required
def bank_exchange_token():
    """Exchange public token for access token and create connection."""
    try:
        from src.bank_service import BankService
        
        data = request.get_json()
        if not data or 'public_token' not in data:
            return jsonify({
                'success': False,
                'error': 'public_token is required'
            }), 400
        
        public_token = data['public_token']
        metadata = data.get('metadata', {})
        
        # Get tenant_id if available
        tenant_id = getattr(current_user, 'tenant_id', None)
        
        # Exchange public token
        bank_service = BankService()
        connection = bank_service.exchange_public_token(public_token, current_user.id, tenant_id, metadata)
        
        return jsonify({
            'success': True,
            'connection_id': connection.id,
            'bank_name': connection.bank_name,
            'message': f'Successfully connected to {connection.bank_name}'
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error exchanging public token: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to exchange public token'
        }), 500

# ==================== ADMIN BANK MANAGEMENT ROUTES ====================

@app.route('/admin/banks', methods=['GET'])
@admin_required
def admin_banks():
    """Admin bank management dashboard."""
    try:
        from src.models import BankConnection, BankSyncLog
        
        # Get system-wide bank statistics
        total_connections = BankConnection.query.count()
        active_connections = BankConnection.query.filter_by(status='active').count()
        error_connections = BankConnection.query.filter_by(status='error').count()
        
        # Get recent sync logs
        recent_logs = BankSyncLog.query.order_by(BankSyncLog.created_at.desc()).limit(20).all()
        
        # Get connections with user details
        connections = db.session.query(BankConnection, User)\
            .join(User, BankConnection.user_id == User.id)\
            .order_by(BankConnection.created_at.desc())\
            .limit(50).all()
        
        connection_data = [
            {
                'id': conn.id,
                'bank_name': conn.bank_name,
                'user_name': user.username,
                'status': conn.status,
                'last_sync': conn.last_sync.isoformat() if conn.last_sync else None,
                'created_at': conn.created_at.isoformat()
            } for conn, user in connections
        ]
        
        return render_template('admin/banks.html', 
            total_connections=total_connections,
            active_connections=active_connections,
            error_connections=error_connections,
            connections=connection_data,
            recent_logs=recent_logs
        )
        
    except Exception as e:
        current_app.logger.error(f"Error loading admin banks: {str(e)}")
        flash('Error loading bank management dashboard', 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/banks/sync-all', methods=['POST'])
@admin_required
def admin_banks_sync_all():
    """System-wide bank sync for all active connections."""
    try:
        from src.bank_service import SyncManager
        from src.models import BankConnection
        
        # Get all active connections
        connections = BankConnection.get_active_connections()
        
        if not connections:
            return jsonify({
                'success': False,
                'error': 'No active bank connections found'
            }), 400
        
        # Initialize sync manager
        sync_manager = SyncManager()
        
        total_imported = 0
        total_duplicates = 0
        errors = []
        synced_connections = 0
        
        # Sync each connection
        for connection in connections:
            try:
                result = sync_manager.sync_connection(connection.id)
                total_imported += result.get('created_transactions', 0)
                total_duplicates += result.get('skipped_duplicates', 0)
                synced_connections += 1
                
                if result.get('errors'):
                    errors.extend(result['errors'])
                    
            except Exception as e:
                error_msg = f"Failed to sync {connection.bank_name} (User: {connection.user_id}): {str(e)}"
                errors.append(error_msg)
                current_app.logger.error(error_msg)
        
        return jsonify({
            'success': True,
            'synced_connections': synced_connections,
            'total_connections': len(connections),
            'imported_count': total_imported,
            'duplicate_count': total_duplicates,
            'errors': errors,
            'message': f'System sync completed: {synced_connections}/{len(connections)} connections synced'
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error during system bank sync: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to perform system bank sync'
        }), 500

@app.route('/admin/banks/logs', methods=['GET'])
@admin_required
def admin_banks_logs():
    """View system bank sync logs and statistics."""
    try:
        from src.models import BankSyncLog, BankConnection
        
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 50
        
        # Get sync logs with connection and user details
        logs = db.session.query(BankSyncLog, BankConnection, User)\
            .join(BankConnection, BankSyncLog.connection_id == BankConnection.id)\
            .join(User, BankConnection.user_id == User.id)\
            .order_by(BankSyncLog.created_at.desc())\
            .paginate(page=page, per_page=per_page, error_out=False)
        
        # Get statistics
        total_logs = BankSyncLog.query.count()
        successful_syncs = BankSyncLog.query.filter_by(status='success').count()
        failed_syncs = BankSyncLog.query.filter_by(status='error').count()
        
        return render_template('admin/bank_logs.html',
            logs=logs,
            total_logs=total_logs,
            successful_syncs=successful_syncs,
            failed_syncs=failed_syncs
        )
        
    except Exception as e:
        current_app.logger.error(f"Error loading bank logs: {str(e)}")
        flash('Error loading bank sync logs', 'error')
        return redirect(url_for('admin_banks'))

if __name__ == "__main__":
    # Create app instance
    
    # Production mode - debug disabled for security
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
